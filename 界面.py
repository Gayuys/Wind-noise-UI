import sys
import os
from PySide6.QtWidgets import QApplication, QFileDialog, QLabel, QWidget, QMessageBox,QFrame, QStyleOption, QStyledItemDelegate, QMainWindow, QVBoxLayout, QTreeWidget, QComboBox, QTableWidget, QTableWidgetItem
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QApplication, QWidget, QMessageBox
from PySide6.QtCore import QFile, Qt
from PySide6.QtGui import QPixmap, QImage,QPainter
from PySide6.QtCore import Qt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
import trimesh
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from typing import Tuple
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import shutil
import pandas as pd
import typing
from PySide6.QtCore import QSize,QTimer
#相关程序导入
import xcepxin_train
import MIV_calculate
import model_use
import optimization_pinduan
import optimization_xiangdu
import Objective_Definition
import characteristics_rating
import noise_rating

# 设置 Matplotlib 中文字体，解决中文显示问题
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimSun', 'Arial']  # 优先使用支持中文的字体
plt.rcParams['axes.unicode_minus'] = False  # 确保负号正确显示
current_dir = os.path.dirname(os.path.abspath(__file__)) #获取当前程序所在文件夹


def load_stl_and_plot_separate_views(stl_path):
    try:
        mesh = trimesh.load_mesh(stl_path)
        vertices = mesh.vertices
        print(f"STL文件加载成功！顶点数：{len(vertices)}，面数：{len(mesh.faces)}")
    except FileNotFoundError:
        print(f"错误：未找到STL文件，请检查路径：{stl_path}")
        return None
    except Exception as e:
        print(f"加载STL文件失败：{str(e)}")
        return None

    separate_views = [
        {"x_coord": vertices[:, 0], "y_coord": vertices[:, 2],
         "plot_title": "正视图（X-Z平面投影）", "x_label": "X轴", "y_label": "Z轴", "window_title": "正视图"},
        {"x_coord": vertices[:, 0], "y_coord": vertices[:, 1],
         "plot_title": "俯视图（X-Y平面投影）", "x_label": "X轴", "y_label": "Y轴", "window_title": "俯视图"},
        {"x_coord": vertices[:, 1], "y_coord": vertices[:, 2],
         "plot_title": "侧视图（Y-Z平面投影）", "x_label": "Y轴", "y_label": "Z轴", "window_title": "侧视图"}
    ]

    pixmaps = []
    point_size = 2
    for view in separate_views:
        fig = plt.figure(figsize=(4, 3), dpi=100)  # 调整大小以适应 QLabel
        plt.scatter(view["x_coord"], view["y_coord"], color='g', s=point_size, alpha=0.7, label="模型顶点")
        plt.title(view["plot_title"], fontsize=10, fontweight='bold', pad=10)
        plt.xlabel(view["x_label"], fontsize=8)
        plt.ylabel(view["y_label"], fontsize=8)
        plt.axis('equal')
        plt.grid(True, alpha=0.3, linestyle='--')
        plt.legend(fontsize=8)
        plt.tight_layout()

        # 将 Matplotlib 图形转换为 QPixmap
        buf = BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        image = QImage.fromData(buf.getvalue())
        pixmap = QPixmap.fromImage(image)
        pixmaps.append(pixmap)
        plt.close(fig)  # 关闭图形以释放内存
        buf.close()

    return pixmaps  # 返回三个视图的 QPixmap 列表

def degrees_to_radians(angles: Tuple[float, float, float]) -> Tuple[float, float, float]:
    """将角度（度）转换为弧度"""
    return tuple(np.radians(angle) for angle in angles)

def create_rotation_matrices(rx: float, ry: float, rz: float) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """创建绕X、Y、Z轴的旋转矩阵"""
    R_x = np.array([
        [1, 0, 0],
        [0, np.cos(rx), -np.sin(rx)],
        [0, np.sin(rx), np.cos(rx)]
    ])
    R_y = np.array([
        [np.cos(ry), 0, np.sin(ry)],
        [0, 1, 0],
        [-np.sin(ry), 0, np.cos(ry)]
    ])
    R_z = np.array([
        [np.cos(rz), -np.sin(rz), 0],
        [np.sin(rz), np.cos(rz), 0],
        [0, 0, 1]
    ])
    return R_x, R_y, R_z

def rotate_stl_vertices(vertices: np.ndarray, rx: float, ry: float, rz: float,
                        rotation_order: str = "xyz") -> np.ndarray:
    """对STL模型的顶点进行绕轴旋转"""
    center = np.mean(vertices, axis=0)
    vertices_centered = vertices - center
    rx_rad, ry_rad, rz_rad = degrees_to_radians((rx, ry, rz))
    R_x, R_y, R_z = create_rotation_matrices(rx_rad, ry_rad, rz_rad)

    rotation_matrix = np.eye(3)
    for axis in rotation_order.lower():
        if axis == "x":
            rotation_matrix = rotation_matrix @ R_x
        elif axis == "y":
            rotation_matrix = rotation_matrix @ R_y
        elif axis == "z":
            rotation_matrix = rotation_matrix @ R_z
        else:
            raise ValueError(f"无效的旋转轴：{axis}，仅支持'x'、'y'、'z'")

    vertices_rotated = vertices_centered @ rotation_matrix.T
    vertices_final = vertices_rotated + center
    return vertices_final

def create_rotated_stl(mesh: trimesh.Trimesh, rotated_vertices: np.ndarray) -> trimesh.Trimesh:
    """基于旋转后的顶点创建新的STL网格对象"""
    rotated_mesh = trimesh.Trimesh(
        vertices=rotated_vertices,
        faces=mesh.faces,
        metadata=mesh.metadata
    )
    return rotated_mesh

def plot_rotated_views(rotated_mesh: trimesh.Trimesh, rx: float, ry: float, rz: float):
    """绘制旋转后模型的三视图，并返回 QPixmap 列表"""
    rot_verts = rotated_mesh.vertices
    views = [
        {"title": f"旋转后正视图（X-Z）\n(绕X:{rx}° Y:{ry}° Z:{rz}°)", "x": rot_verts[:, 0], "y": rot_verts[:, 2],
         "x_label": "X轴", "y_label": "Z轴"},
        {"title": f"旋转后俯视图（X-Y）\n(绕X:{rx}° Y:{ry}° Z:{rz}°)", "x": rot_verts[:, 0], "y": rot_verts[:, 1],
         "x_label": "X轴", "y_label": "Y轴"},
        {"title": f"旋转后侧视图（Y-Z）\n(绕X:{rx}° Y:{ry}° Z:{rz}°)", "x": rot_verts[:, 1], "y": rot_verts[:, 2],
         "x_label": "Y轴", "y_label": "Z轴"}
    ]

    pixmaps = []
    for view in views:
        fig = plt.figure(figsize=(4, 3), dpi=100)
        plt.scatter(view["x"], view["y"], c='crimson', s=1, alpha=0.6, label="旋转后模型")
        plt.title(view["title"], fontsize=10, fontweight='bold')
        plt.xlabel(view["x_label"], fontsize=8)
        plt.ylabel(view["y_label"], fontsize=8)
        plt.axis('equal')
        plt.grid(True, alpha=0.3)
        plt.legend(fontsize=8)
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        image = QImage.fromData(buf.getvalue())
        pixmap = QPixmap.fromImage(image)
        pixmaps.append(pixmap)
        plt.close(fig)
        buf.close()

    return pixmaps

# ---------------- 主窗口类 ---------------- #
class BackgroundFrame(QFrame):
    def __init__(self, parent=None, bg_image_path=None):
        super().__init__(parent)
        self.bg_pixmap = QPixmap()
        if bg_image_path and os.path.exists(bg_image_path):
            self.bg_pixmap = QPixmap(bg_image_path)
        else:
            print(f"警告：背景图路径无效或文件不存在：{bg_image_path}")

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        if not self.bg_pixmap.isNull():
            rect = self.rect()
            # 🔴 改成“按比例扩展铺满控件（允许裁剪）”
            scaled_pixmap = self.bg_pixmap.scaled(
                rect.size(),
                Qt.KeepAspectRatioByExpanding,  # 扩展到覆盖整个控件
                Qt.SmoothTransformation
            )
            # 居中裁剪显示
            pixmap_rect = scaled_pixmap.rect()
            pixmap_rect.moveCenter(rect.center())
            painter.drawPixmap(rect, scaled_pixmap, pixmap_rect)  # 用控件区域裁剪图片

        super().paintEvent(event)

class MyWindow:
    def __init__(self):
        # 1.加载登录界面
        login_window_name = "login.ui"  # 登录界面ui文件
        login_window_file = os.path.join(current_dir, login_window_name)
        self.current_window = self.load_ui(login_window_file)
        if not self.current_window:
            return
        # 2. 替换背景QFrame（必须修改这里的objectName！）
        TARGET_FRAME_NAME = "frame_2"  # 🔴 改成你Qt Designer中背景QFrame的objectName（比如frame、frame_1）
        original_frame = self.current_window.findChild(QFrame, TARGET_FRAME_NAME)
        if not original_frame:
            print(f"错误：找不到名为'{TARGET_FRAME_NAME}'的QFrame，请检查objectName！")
            return

        # 3. 手动指定背景图路径（避免解析样式表的问题，直接写绝对路径）
        bg_folder_name = "绘图\登录背景"  # 背景图所在文件夹（单独文件夹，不要包含文件名）
        bg_image_name = "登录背景.png"  # 背景图文件名
        bg_image_path = os.path.join(current_dir, bg_folder_name, bg_image_name)  # 正确拼接路径

        # 检查路径是否有效
        if not os.path.exists(bg_image_path):
            print(f"错误：背景图文件不存在！路径：{bg_image_path}")
            return

        # 4. 创建自定义Frame并替换（修改这部分）
        parent_widget = original_frame.parentWidget()
        layout = original_frame.layout()

        self.custom_frame = BackgroundFrame(parent=parent_widget, bg_image_path=bg_image_path)
        self.custom_frame.setObjectName(original_frame.objectName())
        self.custom_frame.setStyleSheet(original_frame.styleSheet())

        # 🔴 移除setGeometry，改用布局约束（让Frame随父控件自适应）
        if parent_widget.layout():
            parent_widget.layout().replaceWidget(original_frame, self.custom_frame)
        else:
            # 若父控件无布局，设置Frame为父控件的中心部件
            parent_widget.setCentralWidget(self.custom_frame)

        # 转移布局（保留子控件）
        if layout:
            self.custom_frame.setLayout(layout)

            # 显示自定义Frame，隐藏原Frame
            original_frame.hide()
            self.custom_frame.show()

            # 显示窗口
            self.current_window.show()

        # 绑定登录按钮（你 UI 中的 pushButton）
        if hasattr(self.current_window, "pushButton"):
            self.current_window.pushButton.clicked.connect(self.handle_login_button)
        else:
            print("⚠️ 警告：login.ui 中未找到 pushButton 组件")

        self.current_window.show()
        
        #定义用于数据交换所需的全局变量
        self.model_file = "./缓存/best_model.pth"  # 要移动的模型路径
        self.histroy_data = "./缓存/histroy_data.xlsx"  # 要移动的历史数据路径
        self.input_file_path = None
        self.output_file_path = None
        self.lingmingdujieguo_path = None
        self.mubiaoquxian = None
        self.zaoxingfanwei = "./缓存/目标定义范围.xlsx"
        self.Characteristic_name = "./data/灵敏度分析特征.xlsx"
        self.all_characteristic = "./data/造型优化特征.xlsx"
        self.zaoxingquanzhong = "./data/权重系数表.xlsx"
        self.huancun = "./缓存"
        # self.model_file = "./bestmodel.pth"  # 要移动的模型路径
        # self.model_file = "./bestmodel.pth"  # 要移动的模型路径
    def change_page(self, current_item, previous_item):
        """树形控件切换堆叠页面核心方法"""
        if not current_item:
            return

        click_text = current_item.text(0)
        # 页面映射字典（根据实际UI的stackedWidget页面索引调整）
        page_dict = {
            "⚙️参数设置": 0,
            "📌目标定义": 1,
            "STL模型预处理": 2,
            "造型特征参数提取": 3,
            "造型符合度评分": 4,
            "📈预测模型": 5,
            "灵敏度分析": 6,
            "基于具体频段为目标": 7,
            "基于噪声响度为目标": 8
        }

        if click_text in page_dict:
            target_page_index = page_dict[click_text]
            if hasattr(self.current_window, "stackedWidget"):
                self.current_window.stackedWidget.setCurrentIndex(target_page_index)
                print(f"✅ 切换到页面: {click_text} (索引: {target_page_index})")
            else:
                QMessageBox.warning(self, "警告", "主界面未找到stackedWidget控件！")
        else:
            print(f"ℹ️ 点击了 '{click_text}'，未配置对应页面")

    def switch_to_main_ui(self):
        """切换到主界面 UIzhujiemian.ui"""
        # 关闭当前窗口
        if self.current_window:
            self.current_window.close()

        # 加载新的主界面 UI
        zhujiemian_window_name = "UIzhujiemian.ui" #主界面ui文件
        zhujiemian_window_file = os.path.join(current_dir, zhujiemian_window_name)
        self.current_window = self.load_ui(zhujiemian_window_file)
        if not self.current_window:
            return

        # ─────────────── 重要 ───────────────
        # 找到你的 QTreeWidget（请确认 objectName 是否真的是 treeWidget）
        tree = self.current_window.findChild(QTreeWidget, "treeWidget")   # ← 改成你实际的名称！
        
        if tree:
            # 断开旧的（避免重复绑定）
            try:
                tree.currentItemChanged.disconnect()
            except TypeError:
                pass
            
            # 绑定我们刚写的槽函数
            tree.currentItemChanged.connect(self.change_page)
            print("树形控件切换页面信号已绑定")
        else:
            print("警告：主界面中找不到名为 treeWidget 的 QTreeWidget")
        # ─────────────────────────────────────

        # ←←← 新增：主界面加载完毕后，自动加载14张示意图
        QTimer.singleShot(100, self.load_styling_schematic_images)

        self.current_window.show()

    def check_login_valid(self) -> bool:
        """验证登录账号和密码"""
        user = self.current_window.lineEdit_1.text().strip() if hasattr(self.current_window, "lineEdit_1") else ""
        password = self.current_window.lineEdit_2.text().strip() if hasattr(self.current_window, "lineEdit_2") else ""

        if user == "Faw" and password == "19530715":
            return True
        else:
            QMessageBox.warning(self.current_window, "登录失败", "账号或密码错误，请重新输入！")
            return False
    #绑定登录按钮点击事件
    def handle_login_button(self):
        """点击登录按钮后执行登录验证并跳转主界面"""
        if self.check_login_valid():
           self.switch_to_main_ui()

        # ---------------- 参数设置模块功能按钮 ---------------- #
        #------模型训练功能---------
        # 选择 目标定义数据集
        if hasattr(self.current_window, "pushButton_32"):
            self.current_window.pushButton_32.clicked.connect(self.select_file_yucemoxing_input)
        # 输出 目标定义结果
        if hasattr(self.current_window, "pushButton_36"):
            self.current_window.pushButton_36.clicked.connect(self.select_file_yucemoxing_output)
        if hasattr(self.current_window, "CPB_1"):
            self.current_window.CPB_1.clicked.connect(self.model_train)
        #保存模型
        if hasattr(self.current_window, "CPB_2"):
            self.current_window.CPB_2.clicked.connect(self.save_model)
        # 加载模型
        if hasattr(self.current_window, "CPB_3"):
            self.current_window.CPB_3.clicked.connect(self.select_Data_folder_canshushezhi)

        # ---------------- 目标定义模块功能按钮 ---------------- #
        # 选择主驾驶噪声
        if hasattr(self.current_window, "MPB_1"):
            self.current_window.MPB_1.clicked.connect(self.select_Data_file)
        # 输入 预测模型
        if hasattr(self.current_window, "MPB_3"):
            self.current_window.MPB_3.clicked.connect(self.select_Data_folder_mubiaodingyi)
        # 输出结果
        if hasattr(self.current_window, "MPB_2"):
            self.current_window.MPB_2.clicked.connect(self.mubiaodingyi_result)
        #绘制方案曲线
        if hasattr(self.current_window, "MPB_4"):
            self.current_window.MPB_4.clicked.connect(self.plot_mubiaodingyi_result)       

        # ---------------- 造型评估模块功能按钮 ---------------- #
        #STL文件预处理
        # 选择 STL 文件
        if hasattr(self.current_window, "ZSPB_1"):
            self.current_window.ZSPB_1.clicked.connect(self.select_file)
        # 显示原始三视图
        if hasattr(self.current_window, "ZSPB_2"):
            self.current_window.ZSPB_2.clicked.connect(self.run_stl_plot)
        # 执行旋转并显示旋转后三视图
        if hasattr(self.current_window, "ZSPB_3"):
            self.current_window.ZSPB_3.clicked.connect(self.run_stl_rotation)
        # 选择保存路径
        if hasattr(self.current_window, "ZSPB_4"):
            self.current_window.ZSPB_4.clicked.connect(self.save_rotated_stl)
        
        # 造型提取
        if hasattr(self.current_window, "pushButton_17"):
            self.current_window.pushButton_17.clicked.connect(self.select_file_2)
        # 点击 pushButton_8 输入数据（车高计算、SUV/轿车数据填充）
        if hasattr(self.current_window, "pushButton_18"):
            self.current_window.pushButton_18.clicked.connect(self.run_height_and_fill_data)
        if hasattr(self.current_window, "pushButton_19"):
            self.current_window.pushButton_19.clicked.connect(self.fill_default_values)

        #------造型符合度功能按钮---------
        #导入造型参数值
        if hasattr(self.current_window, "ZCSPB_1"):
            self.current_window.ZCSPB_1.clicked.connect(self.select_zaoxingpingu_zaoxingdaoru_file)
        #导入预测模型
        if hasattr(self.current_window, "ZCSPB_2"):
            self.current_window.ZCSPB_2.clicked.connect(self.select_Data_folder_zaoxingpingu)            
        #计算造型符合度
        if hasattr(self.current_window, "ZCSPB_3"):
            self.current_window.ZCSPB_3.clicked.connect(self.calculate_zaoxingpingu)
              
           


        # ---------------- 预测模型模块功能按钮 ---------------- #
        #------模型预测---------
        #导入模型
        if hasattr(self.current_window, "YPB_1"):
            self.current_window.YPB_1.clicked.connect(self.select_folder_yucemoxing_model)
        #导入预测值
        if hasattr(self.current_window, "YPB_2"):
            self.current_window.YPB_2.clicked.connect(self.select_file_yucemoxing_predict)
        #执行模型预测
        if hasattr(self.current_window, "YPB_3"):
            self.current_window.YPB_3.clicked.connect(self.plot_photo_moxingyuce)
        #保存预测结果
        if hasattr(self.current_window, "YPB_4"):
            self.current_window.YPB_4.clicked.connect(self.save_moxingyuce_result)


        #---------------- 造型优化模块功能按钮 ---------------- #
        
        #------灵敏度分析功能按钮---------
        #点击导入模型及数据集
        if hasattr(self.current_window, "ZLPB_1"):
            self.current_window.ZLPB_1.clicked.connect(self.select_folder_lingmingdu)
        #点击导入数据
        if hasattr(self.current_window, "ZLPB_2"):
            self.current_window.ZLPB_2.clicked.connect(self.select_lingmingduData_file)
        #点击进行灵敏度分析
        if hasattr(self.current_window, "ZLPB_3"):
            self.current_window.ZLPB_3.clicked.connect(self.MIV_Analysis)
        #结果保存
        if hasattr(self.current_window, "ZLPB_4"):
            self.current_window.ZLPB_4.clicked.connect(self.save_results)
        
        #----基于具体频段-----
        if hasattr(self.current_window, "ZJPPB_1"):
            self.current_window.ZJPPB_1.clicked.connect(self.select_folder_pinduan)
        if hasattr(self.current_window, "ZJPPB_2"):
            self.current_window.ZJPPB_2.clicked.connect(self.select_file_zxyh_pinduan)
        if hasattr(self.current_window, "ZJPPB_3"):
            self.current_window.ZJPPB_3.clicked.connect(self.plot_moxingyouhua_pinduan)  #运行优化
        if hasattr(self.current_window, "ZJPPB_4"):
            self.current_window.ZJPPB_4.clicked.connect(self.save_result_pinduan)  # 保存优化结果
   
        #----基于整体响度-----
        if hasattr(self.current_window, "ZJXPB_1"):
            self.current_window.ZJXPB_1.clicked.connect(self.select_folder_xiangdu)
        if hasattr(self.current_window, "ZJXPB_2"):
            self.current_window.ZJXPB_2.clicked.connect(self.select_file_zxyh_xiangdu)
        if hasattr(self.current_window, "ZJXPB_3"):
            self.current_window.ZJXPB_3.clicked.connect(self.plot_moxingyouhua_xiangdu) #运行优化
        if hasattr(self.current_window, "ZJXPB_4"):
            self.current_window.ZJXPB_4.clicked.connect(self.save_result_xiangdu)  # 保存优化结果


        # 显示主界面
        self.current_window.show()

    # ---------------- 登陆界面模块功能 ---------------- #
    def load_ui(self, path):
        ui_file = QFile(path)
        if not ui_file.open(QFile.ReadOnly):
            print(f"❌ 无法打开UI文件: {ui_file.errorString()}")
            return None
        loader = QUiLoader()
        window = loader.load(ui_file)
        ui_file.close()
        if not window:
            print(f"❌ UI加载失败: {loader.errorString()}")
            return None
        return window

    # ---------------- 参数设置模块功能 ---------------- #
    
    #----模型训练------
    #读取造型及技术方案   
    def select_file_yucemoxing_input(self):
        file_path, _ = QFileDialog.getOpenFileName(
        self.current_window,
        "选择文件",
        "",
        "造型及技术方案文件 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "C_1"):
            self.current_window.C_1.setText(file_path)
            
     #导入造型数据库
    def select_file_yucemoxing_output(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window,
            "选择文件",
            "",
            "车内噪声文件 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "C_2"):
            self.current_window.C_2.setText(file_path)  
    #绘制箱型图
    def plot_fitness_history(self, best_fitness_history, avg_fitness_history,max_generations,widget_name):
        """绘制最佳适应度和平均适应度的折线图，并根据QWidget的尺寸调整图像大小"""
        
        # 获取 QWidget 的尺寸
        plot_widget = self.current_window.findChild(QWidget, widget_name)
        if not plot_widget:
            print(f"警告: 找不到名为'{widget_name}'的QWidget")
            return
        
        # 获取 QWidget 的宽度和高度
        widget_width = plot_widget.width()
        widget_height = plot_widget.height()
        
        # 创建matplotlib图形并调整图像大小以适应 QWidget_1 的尺寸
        fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))  # 转换为英寸（1英寸=100像素）
        ax.plot(range(1, max_generations + 1), best_fitness_history, 'r-', linewidth=2,
                label='每代最优适应度')
        ax.plot(range(1, max_generations + 1), avg_fitness_history, 'b--', linewidth=2,
                label='每代平均适应度')
        
        # 设置标题和标签
        ax.set_title('遗传算法优化过程中的适应度曲线', fontsize=14)
        ax.set_xlabel('迭代次数', fontsize=14)
        ax.set_ylabel('适应度值', fontsize=14)
        ax.legend()
        
        # 将图表嵌入到 QWidget_2 中
        canvas = FigureCanvas(fig)
        canvas.setParent(plot_widget)
        
        # 清理旧的 canvas（防止重复添加）
        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)
            plot_widget.setLayout(layout)

        # 删除旧的 FigureCanvas
        for i in reversed(range(layout.count())):
            item = layout.itemAt(i)
            widget = item.widget()
            if widget and isinstance(widget, FigureCanvas):
                widget.deleteLater()
        #添加新的 canvas
        layout.addWidget(canvas)
        canvas.draw()

        # 自动适应 QWidget_1 的大小
        canvas.setGeometry(plot_widget.rect())  # 根据QWidget_1的大小来设置图像尺寸
        canvas.setSizePolicy(plot_widget.sizePolicy())
        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)  #设置布局管理器

        # 显示图表
        plot_widget.layout().addWidget(canvas)    
        
    def plot_loss_history(self, losses, val_losses,widget_name):
        """绘制损失函数图，并根据QWidget的尺寸调整图像大小"""
    
        # 获取 QWidget 的尺寸
        plot_widget = self.current_window.findChild(QWidget, widget_name)
        if not plot_widget:
            print(f"警告: 找不到名为'{widget_name}'的QWidget")
            return
        
        # 获取 QWidget 的宽度和高度
        widget_width = plot_widget.width()
        widget_height = plot_widget.height()
        
        # 创建matplotlib图形并调整图像大小以适应 QWidget_1 的尺寸
        fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))  # 转换为英寸（1英寸=100像素）
        ax.plot(range(1, len(losses) + 1), losses, label='训练损失', linewidth=2)
        ax.plot(range(1, len(val_losses) + 1), val_losses, label='验证损失', linewidth=2)
        
        # 设置标题和标签
        ax.set_title('训练+验证损失曲线', fontsize=14)
        ax.set_xlabel('轮次', fontsize=14)
        ax.set_ylabel('损失', fontsize=14)
        ax.legend()
        
        # 将图表嵌入到 QWidget_1 中
        canvas = FigureCanvas(fig)
        canvas.setParent(plot_widget)
        # 清理旧的 canvas（防止重复添加）
        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)
            plot_widget.setLayout(layout)

        # 删除旧的 FigureCanvas
        for i in reversed(range(layout.count())):
            item = layout.itemAt(i)
            widget = item.widget()
            if widget and isinstance(widget, FigureCanvas):
                widget.deleteLater()
        #添加新的 canvas
        layout.addWidget(canvas)
        canvas.draw()

        # 自动适应 QWidget_1 的大小
        canvas.setGeometry(plot_widget.rect())  # 根据QWidget_1的大小来设置图像尺寸
        canvas.setSizePolicy(plot_widget.sizePolicy())
        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)  #设置布局管理器

        # 显示图表
        plot_widget.layout().addWidget(canvas)
        
    def plot_boxplot(self, errors,widget_name):
        """绘制 losses 和 val_losses 之间差值的箱型图"""
        #定义频率刻度
        frequencies = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                   2000, 2500, 3150, 4000, 5000, 6300, 8000]
        
        # 获取 QWidget_1 的尺寸
        plot_widget = self.current_window.findChild(QWidget, widget_name)
        if not plot_widget:
            print(f"警告: 找不到名为'{widget_name}'的QWidget")
            return
        
        # 获取 QWidget_1 的宽度和高度
        widget_width = plot_widget.width()
        widget_height = plot_widget.height()
        
        # 创建 matplotlib 图形并调整图像大小以适应 QWidget_1 的尺寸
        fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))  # 转换为英寸（1英寸=100像素）

        # 绘制箱型图
        bp = ax.boxplot(
            [errors[:, i] for i in range(errors.shape[1])],
            patch_artist=True,
            widths=0.6,
            tick_labels=[f'{f}Hz' for f in frequencies]
        )

        # 美化颜色
        for patch in bp['boxes']:
            patch.set_facecolor('#A0D8EF')       # 浅蓝色填充
        for whisker in bp['whiskers']:
            whisker.set_color('#333333')
        for cap in bp['caps']:
            cap.set_color('#333333')
        for median in bp['medians']:
            median.set_color('#FF4500')          # 中位数用橙红色
            median.set_linewidth(2)
        for flier in bp['fliers']:
            flier.set(marker='o', color='#FF0000', alpha=0.5, markersize=5)

        # 8. 零误差参考线
        ax.axhline(y=0, color='red', linestyle='--', linewidth=1.5, alpha=0.8, label='零误差线')

        # 9. 坐标轴设置
        ax.set_xlabel('频率 (Hz)', fontsize=12)
        ax.set_ylabel('误差（预测值 - 真实值） (dB)', fontsize=12)
        ax.set_title('各频率点预测误差分布', fontsize=14, pad=15)
        ax.grid(axis='y', linestyle='--', alpha=0.6)
        ax.legend(fontsize=11)
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor') #标签轴旋转
        plt.tight_layout()
        # 显示图表
        # 将图表嵌入到 QWidget_1 中
        canvas = FigureCanvas(fig)
        canvas.setParent(plot_widget)
        # 清理旧的 canvas（防止重复添加）
        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)
            plot_widget.setLayout(layout)

        # 删除旧的 FigureCanvas
        for i in reversed(range(layout.count())):
            item = layout.itemAt(i)
            widget = item.widget()
            if widget and isinstance(widget, FigureCanvas):
                widget.deleteLater()
        #添加新的 canvas
        layout.addWidget(canvas)
        canvas.draw()

        # 自动适应 QWidget_1 的大小
        canvas.setGeometry(plot_widget.rect())  # 根据QWidget_1的大小来设置图像尺寸
        canvas.setSizePolicy(plot_widget.sizePolicy())

        layout = plot_widget.layout()
        if layout is None:
            layout = QVBoxLayout(plot_widget)  #设置布局管理器
        
        # 显示图表
        plot_widget.layout().addWidget(canvas)
    #执行模型训练    
    def model_train(self):
        #处理参数设置
        try:
            ga_max_generations = int(self.current_window.C_4.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法迭代次数必须为数字！")
        try:
            ga_pop_size = int(self.current_window.C_3.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法方案数量必须为数字！")
            
        best_fitness_history, avg_fitness_history, losses, val_losses, y_true_denorm, y_pred_denorm = xcepxin_train.model_Train_main(
             input_file_path=self.current_window.C_1.text().strip(),
             output_file_path=self.current_window.C_2.text().strip(),
             ga_max_generations=ga_max_generations,
             ga_pop_size=ga_pop_size
        )
        errors = np.subtract(y_pred_denorm, y_true_denorm )
        errors_to_list = errors.tolist() if hasattr(errors, 'tolist') else errors
        #绘制损失函数图
        self.plot_loss_history(losses, val_losses, "Cwidget_1") 
        # 绘制适应度历史图
        self.plot_fitness_history(best_fitness_history, avg_fitness_history, ga_max_generations, "Cwidget_2")     
        #绘制箱型图
        self.plot_boxplot(errors, "Cwidget_3")       
        #保存过程数据
        wb = Workbook()
        # 2. 将过程数据写入sheet
        ws1 = wb.active
        ws1.title = "适应度"  # 第一个sheet命名为“适应度”
        ws1.append(best_fitness_history.tolist() if hasattr(best_fitness_history, 'tolist') else best_fitness_history)
        ws1.append(avg_fitness_history.tolist() if hasattr(avg_fitness_history, 'tolist') else avg_fitness_history)
        ws2 = wb.create_sheet(title="损失")
        ws2.append(losses.tolist() if hasattr(losses, 'tolist') else losses)
        ws2.append(val_losses.tolist() if hasattr(val_losses, 'tolist') else val_losses) 
        ws3 = wb.create_sheet(title="误差")
        for row in errors_to_list:
            ws3.append(row)
        wb.save("./缓存/histroy_data.xlsx")
        print("数据已写入多个sheet，文件保存成功！")
        
    #----模型保存------              
    #保存训练好的模型
    def save_model(self):
        """在训练完成后保存训练结果文件"""
        if not hasattr(self, "model_train"):
            print("❌ 尚未进行模型训练，无法保存！")
            return

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存训练好的模型", "", "文件夹 (*)")
        try:
            # 4. 创建新文件夹（exist_ok=False 避免重名）
            os.makedirs(save_path, exist_ok=False)
        except FileExistsError:
            QMessageBox.critical(None, "错误", f"文件夹「{save_path}」已存在！")
            return
        except Exception as e:
            QMessageBox.critical(None, "错误", f"创建文件夹失败：{str(e)}")
            return

        # 5. 检查要移动的模型是否存在
        if not os.path.exists(self.model_file):
            QMessageBox.critical(None, "错误", f"指定文件「{self.model_file}」不存在！")
            return
        if not os.path.exists(self.histroy_data):
            QMessageBox.critical(None, "错误", f"指定文件「{self.histroy_data}」不存在！")
            return
        input_file_path=self.current_window.C_1.text().strip()
        output_file_path=self.current_window.C_2.text().strip()

        # 6. 拼接文件移动后的新路径
        model_name = os.path.basename(self.model_file) 
        new_model_path = os.path.join(save_path, model_name) #保存模型
        data_name = os.path.basename(self.histroy_data)
        new_data_path = os.path.join(save_path, data_name) #保存历史数据
        input_name = os.path.basename(input_file_path)
        new_input_path = os.path.join(save_path, input_name) #保存输入数据
        output_name = os.path.basename(output_file_path)
        new_output_path = os.path.join(save_path, output_name) #保存输入数据

        try:
            # 7. 移动文件到新文件夹
            shutil.move(self.model_file, new_model_path)
            shutil.move(self.histroy_data, new_data_path)
            shutil.move(input_file_path, new_input_path)
            shutil.move(output_file_path, new_output_path)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"移动文件失败：{str(e)}")
            return

        # 8. 弹窗提示文件保存的路径
        QMessageBox.information(
            None, "成功", f"文件已移动至：\n{save_path}"
        )
    
    #-----模型导入功能---------
    def select_Data_folder_canshushezhi(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""
        input_xlsx_path = ""
        output_xlsx_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
                self.input_file_path = input_xlsx_path
            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path
                self.output_file_path = output_xlsx_path
            elif file_name == "histroy_data.xlsx":
                histroy_data = full_path
        #参数设置界面文件路径展示
        if hasattr(self.current_window, "C_5"):
            self.current_window.C_5.setText(pth_path)
        if hasattr(self.current_window, "C_6"):
            self.current_window.C_6.setText(input_xlsx_path)
        if hasattr(self.current_window, "C_7"):
            self.current_window.C_7.setText(output_xlsx_path)
        #目标定义界面文件路径展示
        if hasattr(self.current_window, "M_2"):
            self.current_window.M_2.setText(pth_path)
        #造型评估界面文件路径展示
        if hasattr(self.current_window, "ZCS_2"):
            self.current_window.ZCS_2.setText(pth_path)
        #灵敏度分析界面文件路径展示
        if hasattr(self.current_window, "ZL_1"):
            self.current_window.ZL_1.setText(pth_path)
        #模型预测界面文件路径展示
        if hasattr(self.current_window, "Y_1"):
            self.current_window.Y_1.setText(pth_path)
        #造型优化界面文件路径展示
        if hasattr(self.current_window, "ZJP_1"):
            self.current_window.ZJP_1.setText(pth_path) #基于具体频段优化
        if hasattr(self.current_window, "ZJX_1"):
            self.current_window.ZJX_1.setText(pth_path) #基于具体频段优化

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        msg += f"\n输入数据.xlsx：{input_xlsx_path if input_xlsx_path else '未找到'}"
        msg += f"\n输出数据.xlsx：{output_xlsx_path if output_xlsx_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)
        #解析训练历史数据
        # 读取指定sheet的数据
        df_fitness = pd.read_excel(histroy_data, sheet_name="适应度", header=None)
        df_loss = pd.read_excel(histroy_data, sheet_name="损失", header=None)
        df_error = pd.read_excel(histroy_data, sheet_name="误差", header=None)
        # 读取指定行（pandas的行索引从0开始，与你的原代码逻辑一致）
        best_fitness_history = df_fitness.iloc[0].tolist()  # 第0行=最优适应度
        avg_fitness_history = df_fitness.iloc[1].tolist()  # 第1行=平均适应度
        losses = df_loss.iloc[0].tolist()                  # 第0行=训练损失
        val_losses = df_loss.iloc[1].tolist()              # 第1行=验证损失
        errors = df_error.values.tolist()                  # 读取所有误差数据
        errors = np.array(errors)
        
        #绘制损失函数图
        self.plot_loss_history(losses, val_losses, "Cwidget_4") 
        # 绘制适应度历史图
        ga_max_generations = len(best_fitness_history)
        self.plot_fitness_history(best_fitness_history, avg_fitness_history, ga_max_generations, "Cwidget_5") 
        #绘制箱型图
        self.plot_boxplot(errors, "Cwidget_6") 
        
    # ---------------- 目标定义模块功能 ---------------- #
    def select_Data_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window,
            "选择文件",
            "",
            "竞品车数据 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "M_1"):
            self.current_window.M_1.setText(file_path)
   
    def select_Data_folder_mubiaodingyi(self):
       
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""
        input_xlsx_path = ""
        output_xlsx_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
                self.input_file_path = input_xlsx_path
            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path
                self.output_file_path = output_xlsx_path
        #参数设置界面文件路径展示
        if hasattr(self.current_window, "M_2"):
            self.current_window.M_2.setText(pth_path)
        #灵敏度分析界面文件路径展示
        if hasattr(self.current_window, "ZL_1"):
            self.current_window.ZL_1.setText(pth_path)
        #模型预测界面文件路径展示
        if hasattr(self.current_window, "Y_1"):
            self.current_window.Y_1.setText(pth_path)
        #造型优化界面文件路径展示
        if hasattr(self.current_window, "ZJP_1"):
            self.current_window.ZJP_1.setText(pth_path) #基于具体频段优化
        if hasattr(self.current_window, "ZJX_1"):
            self.current_window.ZJX_1.setText(pth_path) #基于具体频段优化

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        msg += f"\n输入数据.xlsx：{input_xlsx_path if input_xlsx_path else '未找到'}"
        msg += f"\n输出数据.xlsx：{output_xlsx_path if output_xlsx_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)
        
    def mubiaodingyi_result(self):
        """计算参数区间"""    
        try:
            model_path=self.current_window.M_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            original_data_path=self.current_window.M_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择目标车型的数据文件！")        
        
        df_results, feature_names, target_data, top_preds_for_plot = Objective_Definition.make_top10_optimization(model_path, self.input_file_path, self.output_file_path, original_data_path, result_save_path=self.huancun)
        msg = "成功！计算结果已生成"
        QMessageBox.information(None, "计算结果", msg)

    def plot_mubiaodingyi_result(self):
        #绘制符合目标曲线的前十方案
        def plot_top10_vs_target(target_data, top_preds_data, widget_name, title="Top 10 优选方案预测值 vs 目标值对比", save_path=None):
            """
            在指定的UI QWidget中绘制 Top 10 预测方案与目标值的对比折线图
            
            参数:
                target_data: 目标真实值 (1D array/list)
                top_preds_data: Top10 预测值列表，每项为长度与target_data相同的序列
                widget_name: 要嵌入的QWidget的objectName
                title: 图表标题
                save_path: 保存路径（可选），若提供则保存为PNG
            """          
            # 读取 widget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为 '{widget_name}' 的QWidget")
                return
            
            # 获取 widget 当前尺寸（像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建 matplotlib Figure，尺寸大致匹配 widget（dpi≈100）
            fig = Figure(figsize=(widget_width / 100, widget_height / 100))
            ax = fig.add_subplot(111)
            
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 标准1/3倍频程频率点
            std_freqs = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                        2000, 2500, 3150, 4000, 5000, 6300, 8000]
            
            num_points = len(target_data)
            print("target_data:", target_data)
            print("top_preds_data:", top_preds_data)
            print(num_points)
            x_labels = std_freqs[:num_points]
            x_axis = np.arange(num_points)   # 0,1,2,... 用于绘图
            
            # ── 绘制 Top10 预测曲线（灰色细虚线） ──
            colors = ['#003366', '#FF4500', '#228B22', '#8A2BE2', '#00BFFF',
                    '#FF6347', '#556B2F', '#4B0082', '#D2691E', '#00CED1']
            for i in range(len(top_preds_data)):
                # 只给第一条（通常是最优）加图例标签
                label = f'{i+1}号曲线' 
                ax.plot(x_axis, top_preds_data[i],
                        color=colors[i],
                        linestyle='--',
                        linewidth=1,
                        alpha=0.6,
                        label=label)
            
            # ── 高亮 Top1（通常是最优预测） ──
            ax.plot(x_axis, top_preds_data[0],
                    color='#003366',
                    linestyle='--',
                    linewidth=2,
                    label='最佳预测')
            
            # ── 绘制目标真实值（粗红实线） ──
            ax.plot(x_axis, target_data,
                    color='red',
                    marker='o',
                    markersize=6,
                    linewidth=2.5,
                    label='目标数据 (Target)')
            
            # ── 坐标轴设置 ──
            ax.set_xticks(x_axis)
            ax.set_xticklabels(x_labels, rotation=45, ha='right', fontsize=11)
            
            ax.tick_params(axis='y', labelsize=11)
            ax.tick_params(axis='x', pad=8)
            
            ax.set_xlabel('频率 (Hz)', fontsize=13)
            ax.set_ylabel('噪声 (dB)', fontsize=13)
            ax.set_title(title, fontsize=14, pad=15)
            
            ax.legend(fontsize=11, loc='best', framealpha=0.9)
            ax.grid(True, linestyle='--', alpha=0.3)
            
            fig.tight_layout()
            
            # ── 嵌入到 QWidget ──
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            
            # 设置大小策略和几何位置
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 处理布局
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
                layout.setContentsMargins(0, 0, 0, 0)
            
            # 清除旧的 canvas（防止叠加）
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                if item and item.widget() and isinstance(item.widget(), FigureCanvas):
                    item.widget().deleteLater()
            
            layout.addWidget(canvas)
            canvas.draw()
            
            # 可选保存
            if save_path:
                os.makedirs(save_path, exist_ok=True)
                save_file = os.path.join(save_path, "Top10_趋势对比图.png")
                fig.savefig(save_file, dpi=300, bbox_inches='tight')
                print(f"对比图已保存至: {save_file}")
        def plot_comparison(target_data, top_preds_data, widget_name, num_data, title="方案预测值 vs 目标值对比", save_path=None):
            """
            在指定的UI QWidget中绘制 Top 10 预测方案与目标值的对比折线图
            
            参数:
                target_data: 目标真实值 (1D array/list)
                top_preds_data: Top10 预测值列表，每项为长度与target_data相同的序列
                widget_name: 要嵌入的QWidget的objectName
                title: 图表标题
                save_path: 保存路径（可选），若提供则保存为PNG
            """          
            # 读取 widget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为 '{widget_name}' 的QWidget")
                return
            
            # 获取 widget 当前尺寸（像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建 matplotlib Figure，尺寸大致匹配 widget（dpi≈100）
            fig = Figure(figsize=(widget_width / 100, widget_height / 100))
            ax = fig.add_subplot(111)
            
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 标准1/3倍频程频率点
            std_freqs = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                        2000, 2500, 3150, 4000, 5000, 6300, 8000]
            
            num_points = len(target_data)
            x_labels = std_freqs[:num_points]
            x_axis = np.arange(num_points)   # 0,1,2,... 用于绘图
            
            # ── 绘制 Top10 预测曲线（灰色细虚线） ──
            colors = ['#003366', '#FF4500', '#228B22', '#8A2BE2', '#00BFFF',
                    '#FF6347', '#556B2F', '#4B0082', '#D2691E', '#00CED1']
            if num_data == 0 :
                label = '最优方案'
            else:
                label = f'{num_data+2}号曲线'  
            ax.plot(x_axis, top_preds_data[num_data],
                    color=colors[num_data],
                    linestyle='--',
                    linewidth=2.5,
                    alpha=1,
                    label=label)
            
            # ── 绘制目标真实值（粗红实线） ──
            ax.plot(x_axis, target_data,
                    color='red',
                    marker='o',
                    markersize=6,
                    linewidth=2.5,
                    label='目标数据')
            
            # ── 坐标轴设置 ──
            ax.set_xticks(x_axis)
            ax.set_xticklabels(x_labels, rotation=45, ha='right', fontsize=11)
            
            ax.tick_params(axis='y', labelsize=11)
            ax.tick_params(axis='x', pad=8)
            
            ax.set_xlabel('频率 (Hz)', fontsize=13)
            ax.set_ylabel('噪声 (dB)', fontsize=13)
            ax.set_title(title, fontsize=14, pad=15)
            
            ax.legend(fontsize=11, loc='best', framealpha=0.9)
            ax.grid(True, linestyle='--', alpha=0.3)
            
            fig.tight_layout()
            
            # ── 嵌入到 QWidget ──
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            
            # 设置大小策略和几何位置
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 处理布局
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
                layout.setContentsMargins(0, 0, 0, 0)
            
            # 清除旧的 canvas（防止叠加）
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                if item and item.widget() and isinstance(item.widget(), FigureCanvas):
                    item.widget().deleteLater()
            
            layout.addWidget(canvas)
            canvas.draw()
            
            # 可选保存
            if save_path:
                os.makedirs(save_path, exist_ok=True)
                save_file = os.path.join(save_path, f"第{num_data+1}曲线趋势对比图.png")
                fig.savefig(save_file, dpi=300, bbox_inches='tight')
                print(f"对比图已保存至: {save_file}")
        try:
            original_data_path=self.current_window.M_1.text().strip()
            self.mubiaoquxian = self.current_window.M_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择目标车型的数据文件！") 

        preds_data_path='./缓存/Top10_优化方案结果.xlsx'
        #获取需要输出的曲线
        selected_text = self.current_window.comboBox_2.currentText()
        
        target_data = Objective_Definition.load_original_data(original_data_path)  
        top_options = pd.read_excel(preds_data_path, sheet_name='Top10方案', header=0).values
        top_preds_data = pd.read_excel(preds_data_path, sheet_name='Top10_预测结果', header=0).values
        print(top_options)
        
        if selected_text == "全部曲线":
            plot_top10_vs_target(target_data, top_preds_data, 'Mwidget_1', title="Top 10 优选方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[0]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "最优曲线":
            num_data = 0
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "1号曲线":
            num_data = 1
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "2号曲线":
            num_data = 2
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "3号曲线":
            num_data = 3
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "4号曲线":
            num_data = 4
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "5号曲线":
            num_data = 5
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "6号曲线":
            num_data = 6
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "7号曲线":
            num_data = 7
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "8号曲线":
            num_data = 8
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        elif selected_text == "9号曲线":
            num_data = 9
            plot_comparison(target_data, top_preds_data, 'Mwidget_1', num_data, title="方案预测值 vs 目标值对比", save_path=self.huancun)
            data = top_options[num_data]
            for i, value in enumerate(data):
                line_name = f"MN_{i + 1}"
                if hasattr(self.current_window, line_name):
                    formatted_value = f"{value:.3f}"
                    getattr(self.current_window, line_name).setText(formatted_value)
        
    # ---------------- 造型评估模块功能 ---------------- #  
    # ----- STL文件预处理 -----
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window,
            "选择文件",
            "",
            "STL文件 (*.stl);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "ZS_1"):
            self.current_window.ZS_1.setText(file_path)

    def run_stl_plot(self):
        """从 lineEdit 获取 STL 文件路径并将三视图显示在 label_86、label_87、label_88 中"""
        if hasattr(self.current_window, "ZS_1"):
            stl_path = self.current_window.ZS_1.text().strip()
            if stl_path:
                pixmaps = load_stl_and_plot_separate_views(stl_path)
                if pixmaps and len(pixmaps) == 3:
                    if hasattr(self.current_window, "label_86"):
                        self.current_window.label_86.setPixmap(pixmaps[0].scaled(
                            self.current_window.label_86.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
                    else:
                        print("❌ label_86 不存在，请检查 UIXINbuhanbanzidong.ui 文件")
                    if hasattr(self.current_window, "label_87"):
                        self.current_window.label_87.setPixmap(pixmaps[1].scaled(
                            self.current_window.label_87.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
                    else:
                        print("❌ label_87 不存在，请检查 UIXINbuhanbanzidong.ui 文件")
                    if hasattr(self.current_window, "label_88"):
                        self.current_window.label_88.setPixmap(pixmaps[2].scaled(
                            self.current_window.label_88.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
                    else:
                        print("❌ label_88 不存在，请检查 UIXINbuhanbanzidong.ui 文件")
                else:
                    print("❌ 无法生成三视图，请检查 STL 文件！")
            else:
                print("❌ STL文件路径为空，请先选择 STL 文件！")

    def run_stl_rotation(self):
        """执行 STL 旋转并将旋转后三视图显示在 label_95、label_96、label_97 中"""
        if not hasattr(self.current_window, "ZS_1"):
            print("❌ ZS_1 不存在，请检查 UI 文件")
            return

        stl_path = self.current_window.ZS_1.text().strip()
        if not stl_path:
            print("❌  STL文件路径为空，请先选择 STL 文件！")
            return

        # 获取旋转角度
        try:
            rx = float(self.current_window.ZS_2.text().strip()) if hasattr(self.current_window,
                                                                                  "ZS_2") else 0
            ry = float(self.current_window.ZS_3.text().strip()) if hasattr(self.current_window,
                                                                                  "ZS_3") else 0
            rz = float(self.current_window.ZS_4.text().strip()) if hasattr(self.current_window,
                                                                                  "ZS_4") else 0
        except ValueError:
            print("❌ 旋转角度输入无效，请输入有效数字！")
            return

        # 加载 STL 文件
        try:
            self.original_mesh = trimesh.load_mesh(stl_path, force='mesh')
            print(f"原始模型信息：顶点数={len(self.original_mesh.vertices)}，面数={len(self.original_mesh.faces)}")
        except FileNotFoundError:
            print(f"❌ 未找到 STL 文件：{stl_path}")
            return
        except Exception as e:
            print(f"加载 STL 文件失败：{str(e)}")
            return

        # 执行旋转
        print(f"正在执行旋转（顺序：xyz）...")
        self.rotated_vertices = rotate_stl_vertices(
            vertices=self.original_mesh.vertices,
            rx=rx, ry=ry, rz=rz,
            rotation_order="xyz"
        )
        self.rotated_mesh = create_rotated_stl(self.original_mesh, self.rotated_vertices)

        # 生成旋转后三视图并显示
        pixmaps = plot_rotated_views(self.rotated_mesh, rx, ry, rz)
        if pixmaps and len(pixmaps) == 3:
            if hasattr(self.current_window, "label_95"):
                self.current_window.label_95.setPixmap(pixmaps[0].scaled(
                    self.current_window.label_95.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
            if hasattr(self.current_window, "label_96"):
                self.current_window.label_96.setPixmap(pixmaps[1].scaled(
                    self.current_window.label_96.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
            if hasattr(self.current_window, "label_97"):
                self.current_window.label_97.setPixmap(pixmaps[2].scaled(
                    self.current_window.label_97.size(), Qt.IgnoreAspectRatio, Qt.SmoothTransformation))
        else:
            print("❌ 无法生成旋转后三视图，请检查 STL 文件或旋转参数！")

    def save_rotated_stl(self):
        """在旋转完成后保存 STL 文件"""
        if not hasattr(self, "rotated_mesh"):
            print("❌ 尚未旋转 STL，无法保存！")
            return

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存旋转后的 STL 文件", "", "STL Files (*.stl)")
        if save_path:
            try:
                self.rotated_mesh.export(save_path)
                print(f"旋转后的 STL 已保存至：{save_path}")
            except Exception as e:
                print(f"保存旋转后 STL 失败：{str(e)}")

    def select_file_2(self):
        """选择 STL 文件路径，写入 lineEdit_28"""
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window, "选择STL文件", "", "STL文件 (*.stl);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "lineEdit_28"):
            self.current_window.lineEdit_28.setText(file_path)
            print(f"✅ 已选择STL文件：{file_path}")
        else:
            print("❌ 未选择文件或 lineEdit_28 不存在")

    def run_height_and_fill_data(self):
        """计算车高并写入 SUV/轿车数据到 lineEdit_500~548"""
        stl_path = self.current_window.lineEdit_28.text().strip()

        if not stl_path:
            QMessageBox.warning(self.current_window, "提示", "请先选择STL文件！")
            return

        try:
            mesh = trimesh.load_mesh(stl_path)
            vertices = mesh.vertices

            # 计算车高
            z_min = np.min(vertices[:, 2])
            z_max = np.max(vertices[:, 2])
            H = z_max - z_min
            print(f"计算得到车高 H = {H:.2f} mm")

            # SUV 数据
            data1 = [
                "76.41 - 141.75", "26.57 - 63.56", "9.81 - 23.07", "0.07 - 2.89", "6.38 - 8.75",
                "1.76 - 8.24", "5.13 - 20.30", "0.00 - 39.25", "7.14 - 12.46", "75.51 - 126.58",
                "34.06 - 70.15", "5.79 - 32.00", "0.00 - 3.71", "0.00 - 11.58", "4.50 - 12.86",
                "2.42 - 29.03", "0.00 - 45.71", "7.14 - 12.46", "204.01 - 252.34", "209.01 - 250.36",
                "148.94 - 170.74", "63.29 - 87.24", "68.11 - 75.08", "170.72 - 264.00", "17.00 - 22.50",
                "18.00 - 25.00", "149.41 - 157.04", "111.68 - 187.32", "2282.34 - 2876.36", "32.98 - 53.80",
                "38.48 - 65.24", "54.87 - 59.30", "2.60 - 7.74", "22.63 - 42.11", "82.34 - 90.00",
                "1.63 - 2.02", "19 - 22", "21 - 25", "52.10 - 69.81", "37.41 - 73.68",
                "0.00 - 9.48", "2.71 - 3.22","0.85 - 23.04", "33.18 - 60.57", "25.80 - 34.30",
                "78.56 - 81.68", "58.17 - 65.76", "180 - 270", "1.63 - 2.02"
            ]

            # 轿车数据
            data2 = [
                "71.80 - 178.75", "22.17 - 46.09", "2.13 - 41.44", "0.20 - 3.64", "5.97 - 14.98",
                "1.98 - 11.43", "0.19 - 37.75", "0.00 - 29.51", "6.55 - 15.00", "71.42 - 159.29",
                "24.35 - 67.09", "3.53 - 107.43", "0.11 - 3.87", "4.99 - 12.63", "3.64 - 17.77",
                "1.46 - 38.77", "0.00 - 28.04", "6.55 - 15.00", "172.43 - 232.44", "183.01 - 240.44",
                "127.63 - 171.84", "60.66 - 96.24", "69.64 - 77.52", "125.20 - 243.69", "13.00 - 19.00",
                "14.00 - 20.00", "148.97 - 181.66", "8.88 - 148.54", "564.81 - 3244.37", "17.98 - 66.08",
                "12.85 - 79.97", "55.61 - 64.15", "3.19 - 10.07", "12.77 - 59.69", "52.12 - 90.00",
                "1.72 - 2.24", "13 - 18", "14 - 19", "50.46 - 68.26", "40.27 - 69.54",
                "0.00 - 16.54", "2.42 - 3.43","16.14 - 28.41", "28.12 - 89.71", "23.79 - 44.28",
                "75.11 - 84.25", "49.52 - 68.02", "125 - 180", "1.72 - 2.24"
            ]

            # 选择输出数据
            output_data = data1 if H > 1600 else data2
            car_type = "SUV" if H > 1600 else "轿车"
            print(f"检测结果：{car_type}（H = {H:.2f} mm）")

            # 写入 lineEdit_500 ~ lineEdit_548
            for i, value in enumerate(output_data):
                line_name = f"lineEdit_{i + 500}"
                if hasattr(self.current_window, line_name):
                    getattr(self.current_window, line_name).setText(value)

            QMessageBox.information(
                self.current_window,
                "完成",
                f"检测结果：{car_type}\n车高 H = {H:.2f} mm\n数据已写入 lineEdit_500~lineEdit_548"
            )

        except Exception as e:
            QMessageBox.critical(self.current_window, "错误", f"运行出错：\n{e}")

    def fill_default_values(self):
        """点击按钮后向 lineEdit_500~548 写入默认数据（红色）"""

        # SUV 数据
        data1 = [
            "76.41", "26.57", "9.81", "0.07", "6.38",
            "1.76", "5.13", "0.00", "7.14", "75.51",
            "34.06", "5.79", "0.00", "0.00", "4.50",
            "2.42", "0.00", "7.14", "204.01", "209.01",
            "148.94", "63.29", "68.11", "170.72", "17.00",
            "18.00", "149.41", "111.68", "2282.34", "32.98",
            "38.48", "54.87", "2.60", "22.63", "82.34",
            "1.63", "19", "21", "52.10", "37.41",
            "0.00", "2.71", "0.85", "33.18", "25.80",
            "78.56", "58.17", "180", "1.63"
        ]

        # # 轿车数据（如需使用，把 data1 改成 data2 即可）
        # data2 = [
        #     "71.80", "22.17", "2.13", "0.20", "5.97",
        #     "1.98", "0.19", "0.00", "6.55", "71.42",
        #     "24.35", "3.53", "0.11", "4.99", "3.64",
        #     "1.46", "0.00", "6.55", "172.43", "183.01",
        #     "127.63", "60.66", "69.64", "125.20", "13.00",
        #     "14.00", "148.97", "8.88", "564.81", "17.98",
        #     "12.85", "55.61", "3.19", "12.77", "52.12",
        #     "1.72", "13", "14", "50.46", "40.27",
        #     "0.00", "2.42", "16.14", "28.12", "23.79",
        #     "75.11", "49.52", "125", "1.72"
        # ]

        # 选择要填充的数据（默认 SUV）
        values = data1

        # 遍历 lineEdit_500 ~ lineEdit_548
        start_id = 500
        for i, val in enumerate(values):
            obj_name = f"lineEdit_{start_id + i}"

            if hasattr(self.current_window, obj_name):
                le = getattr(self.current_window, obj_name)
                le.setText(val)
                le.setStyleSheet("color: red;")  # 设置红色字体
            else:
                print(f"⚠ 未找到控件：{obj_name}（请检查 UIzhujiemianv3.ui）")

    # --------造型示意图------------
    def load_styling_schematic_images(self):
        """加载14张造型示意图（使用安全的相对路径，兼容直接运行和打包成exe）"""
        # 正确定义根目录
        if getattr(sys, 'frozen', False):
            current_dir = os.path.dirname(sys.executable)
        else:
            current_dir = os.path.dirname(os.path.abspath(__file__))

        folder_name = "绘图/造型示意图"
        folder_path = os.path.join(current_dir, folder_name)

        # 打包后路径兼容
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else current_dir
            folder_path = os.path.join(base_path, folder_name)

        image_names = [
            "A柱上端X向尺寸.png", "A柱上端Y向尺寸.png", "前风挡上端R角.png",
            "A柱下端X向尺寸.png", "A柱下端Y向尺寸.png", "前风挡下端R角.png",
            "后视镜X向尺寸.png", "后视镜Y向尺寸.png", "后视镜末端.png",
            "前轮腔前（后）X向尺寸.png", "后三角窗阶差.png", "顶棚挠度.png",
            "接近角.png", "离去角.png"
        ]

        label_names = [
            "label_16", "label_21", "label_22", "label_27", "label_28",
            "label_40", "label_42", "label_51", "label_53", "label_148",
            "label_61", "label_56", "label_58", "label_59"
        ]

        if len(image_names) != len(label_names):
            print(f"【错误】图片数量({len(image_names)}) ≠ label数量({len(label_names)})")
            return

        success_count = 0

        # 打印表头，方便查看
        print(f"\n{'Label名称':<12} | {'尺寸(WxH)':<12} | {'状态':<8} | {'图片文件'}")
        print("-" * 80)

        for idx, (img_name, label_name) in enumerate(zip(image_names, label_names), 1):
            img_path = os.path.normpath(os.path.join(folder_path, img_name))

            # 1. 检查Label是否存在
            label = self.current_window.findChild(QLabel, label_name)
            if not label:
                print(f"{label_name:<12} | {'None':<12} | ❌ 缺失 | {img_name}")
                continue

            # 2. 【核心修改】在此处获取并打印尺寸
            w, h = label.width(), label.height()

            # 标记尺寸状态
            size_str = f"{w}x{h}"
            # 如果宽或高小于50，通常意味着布局未完成或在隐藏Tab页中，会导致图片缩成点
            status = "⚠️ 极小" if (w < 50 or h < 50) else "✅ 正常"

            print(f"{label_name:<12} | {size_str:<12} | {status} | {img_name}")

            # 3. 检查文件
            if not os.path.exists(img_path):
                print(f"  -> ❌ 图片不存在: {img_path}")
                continue

            pixmap = QPixmap(img_path)
            if pixmap.isNull():
                print(f"  -> ❌ 图片损坏")
                continue

            # 4. 加载图片
            # 【建议】如果发现尺寸极小(status是警告)，强行给一个默认尺寸，防止图片不可见
            target_size = label.size()
            if w < 50 or h < 50:
                # 给一个临时默认值，确保图片能看清（例如 400x300）
                target_size = QSize(699, 536)

            scaled = pixmap.scaled(target_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            label.setPixmap(scaled)
            label.setAlignment(Qt.AlignCenter)
            label.setVisible(True)
            success_count += 1

        if getattr(sys, 'frozen', False):
            print(f"【打包】临时目录路径：{sys._MEIPASS if hasattr(sys, '_MEIPASS') else '未知'}")
        
    #------造型评分功能---------
     #导入造型参数值
    def select_zaoxingpingu_zaoxingdaoru_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window,
            "选择文件",
            "",
            "造型参数文件 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "ZCS_1"):
            self.current_window.ZCS_1.setText(file_path)
            
    #导入预测模型
    def select_Data_folder_zaoxingpingu(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path

        #参数设置界面文件路径展示
        if hasattr(self.current_window, "ZCS_2"):
            self.current_window.ZCS_2.setText(pth_path)
        #灵敏度分析界面文件路径展示
        if hasattr(self.current_window, "ZL_1"):
            self.current_window.ZL_1.setText(pth_path)
        #模型预测界面文件路径展示
        if hasattr(self.current_window, "Y_1"):
            self.current_window.Y_1.setText(pth_path)
        #造型优化界面文件路径展示
        if hasattr(self.current_window, "ZJP_1"):
            self.current_window.ZJP_1.setText(pth_path) #基于具体频段优化
        if hasattr(self.current_window, "ZJX_1"):
            self.current_window.ZJX_1.setText(pth_path) #基于具体频段优化
            
    #显示分析结果
    def calculate_zaoxingpingu(self):
        """计算评价及写入"""

        def plot_pred_vs_target_in_widget(freq_labels, y_pred, y_true, mape, score, widget_name, save_path=None):
            """
            将预测值与目标值的对比图嵌入指定的QWidget中显示
            并可选保存高清图片
            
            参数:
                freq_labels:    频率标签列表，例如 [200, 250, ..., 8000]
                y_pred:         模型预测值 (list/array)
                y_true:         真实目标值 (list/array)
                mape:           MAPE 值（百分比）
                score:          噪声曲线得分
                widget_name:    要嵌入图表的 QWidget 的 objectName
                save_path:      保存图片的路径（含文件名，例如 "result/compare.png"），可选
                current_window: 主窗口实例，用于 findChild（通常是 self 或 self.main_window）
            """
            # 设置中文字体（根据你的环境可能需要调整字体名称）
            plt.rcParams['font.sans-serif'] = ['SimHei', 'STKAITI', 'Microsoft YaHei']
            plt.rcParams['axes.unicode_minus'] = False

            # ------------------- 准备数据 -------------------
            if len(y_pred) != len(y_true) or len(y_true) != len(freq_labels):
                print("警告：y_pred, y_true, freq_labels 长度不一致")
                return

            x_pos = list(range(len(freq_labels)))   # 均匀分布的索引 0,1,2,...,16

            # ------------------- 创建图形 -------------------
            fig = Figure(figsize=(10, 6), dpi=100)   # 初始大小，后续会根据widget调整
            ax = fig.add_subplot(111)

            # 绘制两条折线
            ax.plot(x_pos, y_true, 'ro-', linewidth=2.5, markersize=8, label='目标噪声值')
            ax.plot(x_pos, y_pred, 'bo-', linewidth=2.5, markersize=8, label='模型预测值')

            # 左上角文本框
            ax.text(0.02, 0.98, 
                    f'MAPE = {mape:.2f}% | 噪声曲线得分 = {score:.2f}',
                    transform=ax.transAxes,
                    fontsize=15,
                    verticalalignment='top',
                    bbox=dict(boxstyle='round,pad=0.5', 
                            facecolor='lightgreen', 
                            edgecolor='green',
                            alpha=0.85))

            # 设置坐标轴
            ax.set_xticks(x_pos)
            ax.set_xticklabels(freq_labels, rotation=45, fontsize=11)
            ax.tick_params(axis='y', labelsize=12)
            ax.set_xlabel('频率 (Hz)', fontsize=14, fontweight='bold')
            ax.set_ylabel('噪声值 (dB)', fontsize=14, fontweight='bold')
            ax.set_title('200–8000Hz 噪声值：模型预测 vs 目标值', 
                        fontsize=15, fontweight='bold', pad=15)
            
            ax.legend(fontsize=13, loc='upper right')
            ax.grid(True, alpha=0.3, linestyle='--')

            fig.tight_layout()

            # ------------------- 寻找目标 widget -------------------
            if self.current_window is None:
                print("错误：请传入 current_window 参数（通常是 self 或主窗口）")
                return

            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为 '{widget_name}' 的 QWidget")
                return

            # 获取 widget 当前尺寸（像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()

            # 根据 widget 实际大小动态调整 figure（dpi≈100）
            if widget_width > 100 and widget_height > 100:
                fig.set_size_inches(widget_width / 100, widget_height / 100)

            # 创建 canvas 并嵌入
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)

            # 清理旧的 canvas（防止重复叠加）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
                plot_widget.setLayout(layout)
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setSpacing(0)

            # 删除旧的 FigureCanvas
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                if item and item.widget() and isinstance(item.widget(), FigureCanvas):
                    item.widget().deleteLater()

            # 添加新 canvas 并重绘
            layout.addWidget(canvas)
            canvas.draw()

            # 让 canvas 跟随 widget 大小变化（推荐在主窗口 resizeEvent 中再调用一次 resize）
            canvas.setGeometry(plot_widget.rect())

            # ------------------- 可选保存高清图片 -------------------
            if save_path:
                try:
                    os.makedirs(os.path.dirname(save_path), exist_ok=True)
                    fig.savefig(save_path, dpi=300, bbox_inches='tight')
                    print(f"对比图已保存至：{save_path}")
                except Exception as e:
                    print(f"保存图片失败：{e}")

            # 注意：这里不调用 plt.close()，因为 fig 是我们手动创建的
        def plot_characteristics(col_index, values, tableWidget_name):
            """
            一次性填充指定列
            col_index: 列号（从0开始）
            values:    这列的所有值（列表长度应等于表格当前行数）
            """
            if self.current_window is None:
                print("错误：请传入 current_window 参数（通常是 self 或主窗口）")
                return

            target_table = self.current_window.findChild(QTableWidget, "tableWidget")
            if not target_table:
                print(f"警告: 找不到名为 '{tableWidget_name}' 的 QWidget")
                return
            
            if col_index < 0 or col_index >= target_table.columnCount():
                print(f"列索引 {col_index} 超出范围")
                return

            row_count = target_table.rowCount()
            if len(values) != row_count:
                print(f"数据长度 {len(values)} 与行数 {row_count} 不匹配")
                return

            for row in range(row_count):
                item = QTableWidgetItem(str(values[row]))   # 记得转字符串
                target_table.setItem(row, col_index, item)
                
        try:
            model_path=self.current_window.ZCS_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            new_input_path=self.current_window.ZCS_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择评估车型的数据文件！")
            
        try:
            target_df = pd.read_excel(
                self.mubiaoquxian,
                sheet_name='Sheet1',
                header=None,
                engine='openpyxl'          # 明确用 openpyxl
            )
            print(target_df)
        except Exception as e:
            print("读取参数文件失败：", type(e).__name__)
            print("详细错误：", str(e))
            print("文件路径：", self.mubiaoquxian)
            print("文件是否存在？", os.path.exists(self.mubiaoquxian))
            print("文件大小：", os.path.getsize(self.mubiaoquxian) if os.path.exists(self.mubiaoquxian) else "不存在")
            raise
            
        freq_labels, pred_noise, target_noise, mape, noise_score = noise_rating.calculate_rating(model_path, self.input_file_path, self.output_file_path, new_input_path, self.mubiaoquxian)
        plot_pred_vs_target_in_widget(freq_labels, pred_noise, target_noise, mape, noise_score, 'PGwidget_1', save_path=None)
        shape_score, df_transposed = characteristics_rating.calculate_rating(self.all_characteristic, new_input_path, self.zaoxingfanwei, self.zaoxingquanzhong)
        print(df_transposed)
        print(shape_score)
        origin_data = df_transposed.iloc[0, :24].values
        numeric_data = pd.to_numeric(origin_data, errors='coerce')
        origin_data_rounded = numeric_data.round(2)
        origin_data_list = origin_data_rounded.tolist()
        plot_characteristics(0, origin_data_list, 'tableWidget')
        
        min_data = df_transposed.iloc[1, :24].values
        min_data = np.array(min_data)
        numeric_data = pd.to_numeric(min_data, errors='coerce')
        min_data = np.round(numeric_data, 2)
        min_data_list = min_data.tolist()
        plot_characteristics(1, min_data_list, 'tableWidget')
        
        max_data = df_transposed.iloc[2, :24].values
        max_data = np.array(max_data)
        numeric_data = pd.to_numeric(max_data, errors='coerce')
        max_data_rounded = numeric_data.round(2)
        max_data_list = max_data_rounded.tolist()
        plot_characteristics(2, max_data_list, 'tableWidget')
        
        score_data = df_transposed.iloc[4, :24].values
        score_data = np.array(score_data)
        numeric_data = pd.to_numeric(score_data, errors='coerce')
        score_data_rounded = numeric_data.round(2)
        score_data_list = score_data_rounded.tolist()
        plot_characteristics(3, score_data_list, 'tableWidget')
        
        if hasattr(self.current_window, "ZCS_3"):
            self.current_window.ZCS_3.setText(str(shape_score))
        else:
            print(f"警告: 找不到名为 'ZCS_3' 的 QWidget")
            
        if hasattr(self.current_window, "ZCS_4"):
            self.current_window.ZCS_4.setText(str(mape))
        else:
            print(f"警告: 找不到名为 'ZCS_4' 的 QWidget")
            
        if hasattr(self.current_window, "ZCS_5"):
            self.current_window.ZCS_5.setText(str(noise_score))
        else:
            print(f"警告: 找不到名为 'ZCS_5' 的 QWidget")
        
        try:
            zaoxing_weight = float(self.current_window.ZCS_6.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "造型参数权重必须为数字！")
        try:
            noise_weight = float(self.current_window.ZCS_7.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "噪声权重必须为数字！")
          
        if zaoxing_weight and noise_weight:
            shape_conformity_score = round(shape_score * zaoxing_weight + noise_score * noise_weight, 2)
            if hasattr(self.current_window, "ZCS_8"):
                self.current_window.ZCS_8.setText(str(shape_conformity_score))
            else:
                print(f"警告: 找不到名为 'ZCS_8' 的 QWidget")
        else:
            QMessageBox.warning(self.current_window, "输入错误", "请输入造型参数权重和噪声权重！")
        
        
    # ---------- 预测模型模块功能 ---------------- #    

    #----模型预测------
    #加载模型文件
    def select_folder_yucemoxing_model(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹2")
        if not folder_path:
            return

        pth_path = ""
        input_xlsx_path = ""
        output_xlsx_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path

        if hasattr(self.current_window, "Y_1"):
            self.current_window.Y_1.setText(pth_path)
        # if hasattr(self.current_window, "lineEdit_137"):
        #     self.current_window.lineEdit_137.setText(input_xlsx_path)
        # if hasattr(self.current_window, "lineEdit_115"):
        #     self.current_window.lineEdit_115.setText(output_xlsx_path)

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        msg += f"\n输入数据.xlsx：{input_xlsx_path if input_xlsx_path else '未找到'}"
        msg += f"\n输出数据.xlsx：{output_xlsx_path if output_xlsx_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)
        
    #加载输入数据
    def select_file_yucemoxing_predict(self):
        file_path, _ = QFileDialog.getOpenFileName(
        self.current_window,
        "选择文件",
        "",
        "造型参数+技术方案 (*.xlsx);;所有文件 (*.*)"
    )
        if file_path and hasattr(self.current_window, "Y_2"):
            self.current_window.Y_2.setText(file_path)
            new_input_path=self.current_window.Y_2.text().strip()
            
        new_input_data = pd.read_excel(new_input_path, header=0)
        new_input_data = new_input_data.iloc[0].tolist()
        #new_input_data = np.array(new_input_data).reshape(1, -1)    
        for i, value in enumerate(new_input_data):
            line_name = f"YD_{i + 1}"
            if hasattr(self.current_window, line_name):
                getattr(self.current_window, line_name).setText(str(value))
            
    def plot_photo_moxingyuce(self): 
        
        def visualize_predictions(predicted_data, widget_name, save_path=None, save_path_data=None):
            """
            可视化预测结果，嵌入到指定的 QWidget 中显示
            参数:
                predicted_data: 预测数据列表/数组
                widget_name:    要嵌入图表的 QWidget 的 objectName
                save_path:      保存图片的文件夹路径（可选）
                save_path_data: 保存预测数据的 Excel 路径（可选）
            """
            plt.rcParams['font.sans-serif'] = ['SimHei', 'STKAITI']
            plt.rcParams['axes.unicode_minus'] = False

            min_length = min(len(predicted_data), 17)
            predicted_data_slice = predicted_data[:min_length]

            x_axis_data = [470 * i for i in range(1, min_length + 1)]
            freq_labels = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                        2000, 2500, 3150, 4000, 5000, 6300, 8000][:min_length]

            # ------------------- 创建图形 -------------------
            fig = plt.figure(figsize=(10, 6))  # 先用默认大小，后面会调整

            plt.plot(x_axis_data, predicted_data_slice, color="k", marker='s', linewidth=1.5, label='预测数据')
            plt.xticks(x_axis_data, freq_labels, fontsize=12, rotation=45)
            plt.xlabel('频率(Hz)', fontsize=14)
            plt.ylabel('噪声(dB)', fontsize=14)
            plt.legend(fontsize=12)
            plt.grid(True, alpha=0.3)
            plt.tight_layout()

            # ------------------- 寻找目标 widget -------------------
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为 '{widget_name}' 的 QWidget")
                plt.close(fig)
                return

            # 获取 widget 当前尺寸（像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()

            # 调整 figure 大小以尽量贴合 widget（dpi ≈ 100）
            fig.set_size_inches(widget_width / 100, widget_height / 100)

            # 创建 canvas 并嵌入
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)

            # 清理旧的 canvas（防止重复添加）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
                plot_widget.setLayout(layout)

            # 删除旧的 FigureCanvas
            for i in reversed(range(layout.count())):
                item = layout.itemAt(i)
                widget = item.widget()
                if widget and isinstance(widget, FigureCanvas):
                    widget.deleteLater()

            # 添加新的 canvas
            layout.addWidget(canvas)
            canvas.draw()

            # 让 canvas 跟随 widget 尺寸（推荐配合主窗口的 resizeEvent 使用）
            canvas.setGeometry(plot_widget.rect())

            # 可选：保存图片
            if save_path:
                save_img_path = os.path.join(save_path, '预测结果.png')
                fig.savefig(save_img_path, dpi=300, bbox_inches='tight')
                print(f"图像已保存至: {save_img_path}")

            # 可选：保存数据到 Excel（保持原逻辑）
            if save_path_data:
                import pandas as pd
                df = pd.DataFrame({
                    '频率(Hz)': freq_labels,
                    '预测值': predicted_data_slice
                })
                save_data_path = os.path.join(save_path_data, '预测结果.xlsx')
                df.to_excel(save_data_path, index=False)
                print(f"预测数据已保存至: {save_data_path}")

        try:
            model_path=self.current_window.Y_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            new_input_path=self.current_window.Y_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择进行灵敏度排序的数据文件！")
        #获取造型+技术方案数据
        new_input_data = pd.read_excel(new_input_path, header=0).values
        pride_input_data = pd.read_excel(self.mubiaoquxian, header=0).values
        input_data = []
        for i in range(new_input_data.shape[1]):
            line_name = f"YD_{i + 1}"
            data = float(getattr(self.current_window, line_name).text().strip())
            input_data.append(data)
        #input_data = input_data.iloc[0].tolist()
        print(input_data)
        input_data = np.array(input_data)
        print(input_data)
        print(input_data.shape)
        input_data = input_data.reshape(1, -1)
        input_file_path = self.input_file_path #输入归一化
        output_file_path = self.output_file_path #输出归一化

        predicted_data = model_use.call_model(input_file_path, output_file_path, input_data, model_path)
        visualize_predictions(predicted_data, 'Ywidget', self.huancun, self.huancun)
        # 获取表头
        df_headers = pd.read_excel(new_input_path, header=0)
        column_names = df_headers.columns.tolist()
        # 创建新DataFrame，使用原始表头和input_data作为一行数据
        new_df = pd.DataFrame(input_data, columns=column_names)
        # 保存为Excel文件
        output_path = os.path.join(self.huancun, "造型参数+技术方案.xlsx")  # 或您指定的路径
        new_df.to_excel(output_path, index=False)

    def save_moxingyuce_result(self):
        """在预测完成后保存分析结果"""
        if not hasattr(self, "model_train"):
            print("❌ 尚未进行模型训练，无法保存！")
            return

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存预测结果", "", "文件 (*)")
        try:
            # 4. 创建新文件（exist_ok=False 避免重名）
            os.makedirs(save_path, exist_ok=False)
        except FileExistsError:
            QMessageBox.critical(None, "错误", f"文件夹「{save_path}」已存在！")
            return
        except Exception as e:
            QMessageBox.critical(None, "错误", f"创建文件失败：{str(e)}")
            return

        #设置要移动文件的路径
        input_path = os.path.join(self.huancun, "造型参数+技术方案.xlsx")
        result_path = os.path.join(self.huancun, "预测结果.xlsx")
        photo_path = os.path.join(self.huancun, "预测结果.png")

        # 5. 检查要移动的模型是否存在
        if not os.path.exists(input_path):
            QMessageBox.critical(None, "错误", f"指定文件造型参数+技术方案.xlsx不存在！")
            return
        if not os.path.exists(result_path):
            QMessageBox.critical(None, "错误", f"指定文件预测结果.xlsx不存在！")
            return
        if not os.path.exists(photo_path):
            QMessageBox.critical(None, "错误", f"指定文件预测结果.png不存在！")
            return
        result_name = os.path.basename(result_path) 
        new_result_path = os.path.join(save_path, result_name) #保存预测结果
        photo_name = os.path.basename(photo_path) 
        new_photo_path = os.path.join(save_path, photo_name) #保存展示图片 
        input_name = os.path.basename(input_path) 
        new_input_path = os.path.join(save_path, input_name) #保存输入数据

        try:
            # 7. 移动文件到新文件夹
            shutil.move(result_path, new_result_path)
            shutil.move(photo_path, new_photo_path)
            shutil.move(input_path, new_input_path)
            QMessageBox.critical(None, "完成", f"移动文件成功！\n预测结果已保存至: {new_result_path}\n展示图片已保存至: {new_photo_path}\n输入数据已保存至: {new_input_path}")
        except Exception as e:
            QMessageBox.critical(None, "错误", f"移动文件失败：{str(e)}")
            return     


    # ---------------- 造型优化模块功能 ---------------- #
    
    #----灵敏度分析----
        #--------灵敏度分析功能------------
    def select_folder_lingmingdu(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""
        input_xlsx_path = ""
        output_xlsx_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
                self.input_file_path = input_xlsx_path

            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path
                self.output_file_path = output_xlsx_path

        if hasattr(self.current_window, "ZL_1"):
            self.current_window.ZL_1.setText(pth_path)

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        msg += f"\n输入数据.xlsx：{input_xlsx_path if input_xlsx_path else '未找到'}"
        msg += f"\n输出数据.xlsx：{output_xlsx_path if output_xlsx_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)
        
    def select_lingmingduData_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.current_window,
            "选择文件",
            "",
            "数据集 (*.xlsx);;所有文件 (*.*)"
        )
        if file_path and hasattr(self.current_window, "ZL_2"):
            self.current_window.ZL_2.setText(file_path)

    def MIV_Analysis(self):
        """执行灵敏度分析"""
        def plot_sensitivityonly(miv1, Characteristic_name, y_label, title, widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制灵敏度热力图（heatmap），Y轴显示频率标签，四组数据位于对应位置。
            支持动态适应widget尺寸，并可选保存图像。
            """
            #读取技术方案名称
            file_path = Characteristic_name #获取技术方案名称
            data = pd.read_excel(file_path, header=0)  # 第一行作为列名
            label = data.columns.tolist()
            print(label)
            # 数据处理
            miv = miv1.T
            
            corr_df = pd.DataFrame(miv, index=y_label, columns=label)  # 创建DataFrame用于绘图
            corr_df_reversed = corr_df[::-1]  # 数据取反，使频率从小到大排序（从下到上）
            print(corr_df_reversed)
            
            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（大致1英寸 ≈ 100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 绘制热力图
            sns.heatmap(corr_df_reversed, 
                        annot=True,                     # 显示数值
                        cmap='coolwarm',                # 红蓝配色
                        vmin=np.min(miv), 
                        vmax=np.max(miv), 
                        center=(np.min(miv) + np.max(miv)) / 2,
                        fmt='.2f',                      # 保留2位小数
                        annot_kws={'size': 8},          # 数值字体大小
                        ax=ax,                          # 指定axes
                        cbar_kws={"shrink": 0.8})       # 颜色条调整
            
            # 设置标题和轴标签
            ax.set_title(title, fontsize=14, pad=15)
            ax.set_xlabel('造型特征+技术方案', fontsize=12)
            ax.set_ylabel('频率', fontsize=12)
            
            # 旋转x轴标签，避免重叠
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')
            ax.tick_params(axis='x', which='major', pad=5)
            
            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, f'风噪灵敏度分析结果.png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
            
            # 注意：不需要调用plt.show()或plt.close()，因为嵌入到Qt中由canvas管理
        def plot_sensitivity(miv1, miv2, Characteristic_name, title, widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制灵敏度热力图（heatmap），Y轴显示频率标签，四组数据位于对应位置。
            支持动态适应widget尺寸，并可选保存图像。
            """
        #读取技术方案名称
            file_path = Characteristic_name #获取技术方案名称
            data = pd.read_excel(file_path, header=0)  # 第一行作为列名
            label = data.columns.tolist()
        # 数据处理
            miv1 = miv1.T
            miv2 = miv2.T

            miv = np.vstack([miv1, miv2]) #将列向量转为行向量堆叠
            
            y_positions = ['+10%', '-10%'] #增加值的方向
            
            corr_df = pd.DataFrame(miv, index=y_positions, columns=label) # 创建DataFrame用于绘图
    
            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（大致1英寸 ≈ 100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 绘制热力图
            sns.heatmap(corr_df, 
                        annot=True,                     # 显示数值
                        cmap='coolwarm',                # 红蓝配色
                        vmin=np.min(miv), 
                        vmax=np.max(miv), 
                        center=(np.min(miv) + np.max(miv)) / 2,
                        fmt='.2f',                      # 保留2位小数
                        annot_kws={'size': 8},          # 数值字体大小
                        ax=ax,                          # 指定axes
                        cbar_kws={"shrink": 0.8})       # 颜色条调整
            
            # 设置标题和轴标签
            ax.set_title(title, fontsize=14, pad=15)
            ax.set_xlabel('造型特征+技术方案', fontsize=12)
            ax.set_ylabel('变化范围', fontsize=12)
            
            # 旋转x轴标签，避免重叠
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')
            ax.tick_params(axis='x', which='major', pad=5)
            
            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, f'风噪灵敏度分析结果.png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
            
            # 注意：不需要调用plt.show()或plt.close()，因为嵌入到Qt中由canvas管理  
        def plot_excel_table_widget(excel_path, sheet_name=0, widget_name="table_widget",
                                title="Excel 数据表格预览", save_path=None, dpi=300):
            """
            将 Excel 文件内容以表格形式绘制到指定的 PyQt QWidget 中，
            支持动态适应 widget 尺寸，并可选保存为高质量矢量/位图文件。
            
            参数:
                excel_path: str - Excel 文件路径
                sheet_name: str/int - 要读取的工作表名称或索引，默认第0个
                widget_name: str - 要嵌入图表的 QWidget 对象名称
                title: str - 图表标题
                save_path: str - 保存图片的文件夹路径（可选），None 则不保存
                dpi: int - 保存时的分辨率（仅对光栅格式有效）
            """
                # 1. 读取 Excel 数据（无表头模式，更接近原始表格展示）
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            
            # 删除全空行和全空列
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            if df.empty:
                print("Excel 文件为空或无有效数据")
                return None
                
            data = df.values.tolist()
            n_rows, n_cols = len(data), len(data[0]) if data else 0

            # 2. 查找目标 QWidget
            # 注意：这里假设你有一个全局/类属性 current_window 持有主窗口
            # 如果不是这样，请根据实际情况修改获取 widget 的方式
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为 '{widget_name}' 的 QWidget")
                return None

            # 3. 获取 widget 当前像素尺寸
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()

            # 4. 根据表格大小动态估算图形尺寸（英寸）
            # 每个单元格大约宽度 0.8~1.2 英寸，高度 0.35~0.5 英寸
            cell_width_inch = 1.1
            cell_height_inch = 0.42
            
            fig_width = max(n_cols * cell_width_inch, widget_width / 100)
            fig_height = max(n_rows * cell_height_inch, widget_height / 100)
            
            # 限制最大尺寸，避免过大撑爆界面
            fig_width = min(fig_width, 25)
            fig_height = min(fig_height, 18)

            # 5. 创建 matplotlib Figure
            fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=100)
            ax.axis('off')
            ax.margins(0)

            plt.rcParams['font.sans-serif'] = ['SimHei']  # 支持中文
            plt.rcParams['axes.unicode_minus'] = False
            plt.rcParams['font.size'] = 9

            # 6. 绘制表格
            table = ax.table(
                cellText=data,
                cellLoc='center',
                bbox=[0, 0, 1, 1],
                cellColours=[['#f8f9fa'] * n_cols for _ in range(n_rows)],
                edges='closed'
            )

            # 7. 表格样式优化
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.3, 1.4)  # 可根据实际效果微调

            # 美化表头（第一行）
            for j in range(n_cols):
                if (0, j) in table._cells:
                    cell = table[(0, j)]
                    cell.set_facecolor('#4472c4')
                    cell.set_text_props(color='white', weight='bold')

            # 8. 调整布局，去除多余边距
            plt.subplots_adjust(left=0.01, right=0.99, top=0.96, bottom=0.01)

            # 9. 设置标题（可选显示在图上）
            if title:
                ax.set_title(title, fontsize=13, pad=12)

            # 10. 将图表嵌入到 QWidget 中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()

            # 适应 widget 大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())

            # 处理布局 - 清除旧的 canvas，防止叠加
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
                plot_widget.setLayout(layout)

            # 清理旧的 FigureCanvas
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()

            layout.addWidget(canvas)

            # 11. 可选：保存文件
            saved_file = None
            if save_path:
                os.makedirs(save_path, exist_ok=True)
                
                # 可根据需要选择保存格式
                # 建议保存为矢量格式（svg/pdf）以保持清晰度
                save_file_svg = os.path.join(save_path, "excel_table_preview.svg")
                fig.savefig(save_file_svg, bbox_inches='tight', pad_inches=0.02,
                            facecolor='white', format='svg')
                
                # 也可以同时保存高分辨率 png
                # save_file_png = os.path.join(save_path, "excel_table_preview.png")
                # fig.savefig(save_file_png, dpi=dpi, bbox_inches='tight', facecolor='white')
                
                saved_file = save_file_svg
                print(f"表格已保存至: {saved_file}")

            # 12. 关闭 figure 释放内存（重要！）
            plt.close(fig)
        #输出优化方案
        def sum_and_rank_params_from_heatmap(MIV, param_labels, freq_labels, save_path):
            """
            基于全频段热力图数据，计算每个参数对应的17个频点MIV数据之和，按从大到小排序取前十
            :param MIV: 灵敏度矩阵（形状：参数数×17频点）
            :param param_labels: 参数名称列表（对应热力图的列）
            :param freq_labels: 频点名称列表（对应热力图的行）
            :param save_path: 结果保存路径
            """
            # 1. 构建与热力图一致的DataFrame（参数×频点）
            # 截断参数名称，确保与MIV行数一致
            heatmap_df = pd.DataFrame(MIV,
                                    index=param_labels[:MIV.shape[0]],  # 防止参数名称数量与MIV行数不匹配
                                    columns=freq_labels)  # 列：频点（对应热力图的行）

            # 2. 计算每个参数的17个频点MIV数据之和（对每个参数行求和）
            param_total = heatmap_df.sum(axis=1)  # axis=1：对行求和（每个参数的17个频点）

            # 3. 组合参数名称与对应总和（列名保持一致）
            param_sum_df = pd.DataFrame({
                "参数名称": param_total.index,
                "17频点MIV总和（上调-下调噪声差值绝对值之和）": param_total.values  # 统一列名
            })

            # 4. 按总和【从大到小】排序，取前十名（列名与上面一致，修复KeyError）
            param_sum_sorted = param_sum_df.sort_values(
                by="17频点MIV总和（上调-下调噪声差值绝对值之和）",
                ascending=False  # 改为False，实现从大到小排序
            ).head(10)

            # 5. 打印结果
            print("=" * 80)
            print("全频段热力图-每个参数17个频点MIV总和（从大到小排序，取前十）")
            print("=" * 80)
            print(param_sum_sorted)
            print("=" * 80)

            # 6. 保存到Excel（增加目录创建和异常捕获）
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    param_sum_sorted.to_excel(writer, index=False)
                print(f"✅ 参数MIV总和排序结果已保存到：{save_path}")
            except PermissionError:
                print(f"❌ 权限错误：无法写入 {save_path}，请关闭该文件后重试")
            except Exception as e:
                print(f"❌ 保存失败：{str(e)}")

            return param_sum_sorted
        
        def all_params_from_heatmap(MIV, param_labels, freq_labels, save_path):
            """
            基于全频段热力图数据，计算每个参数对应的17个频点MIV数据之和，按从大到小排序取前十
            :param MIV: 灵敏度矩阵（形状：参数数×17频点）
            :param param_labels: 参数名称列表（对应热力图的列）
            :param freq_labels: 频点名称列表（对应热力图的行）
            :param save_path: 结果保存路径
            """
            # 1. 构建与热力图一致的DataFrame（参数×频点）
            # 截断参数名称，确保与MIV行数一致
            # 确保输入维度匹配
            if MIV.shape[1] != len(freq_labels):
                raise ValueError(f"MIV 列数({MIV.shape[1]}) 与 freq_labels 长度({len(freq_labels)})不匹配")
            if MIV.shape[0] > len(param_labels):
                raise ValueError("param_labels 数量少于 MIV 的行数")
            
            # 截取匹配的参数标签
            param_labels = param_labels[:MIV.shape[0]]
            
            # 构建 DataFrame：行=参数，列=频点
            heatmap_df = pd.DataFrame(
                MIV,
                index=param_labels,
                columns=freq_labels
            )
            
            # 存储每个频点的 Top-10 参数名称
            top10_dict = {}
            
            for freq in freq_labels:
                # 按当前频点的 MIV 值降序排序
                sorted_series = heatmap_df[freq].sort_values(ascending=False)
                
                # 取前10个参数名称
                top10_names = sorted_series.index[:10].tolist()
                
                # 补齐到10个（如果不足10个，用空字符串或 NaN）
                top10_names += [""] * (10 - len(top10_names))
                
                top10_dict[freq] = top10_names
            
            # 构建最终的 (10 × 17) DataFrame
            result_df = pd.DataFrame(top10_dict)
            
            # 可选：设置行索引为 1~10 或 "Rank 1" ~ "Rank 10"
            result_df.index = [f"参数 {i+1}" for i in range(10)]
            
            # 保存到 Excel
            try:
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                result_df.to_excel(save_path, index=True)  # 保留排名索引
                print(f"已保存到：{save_path}")
                print(f"输出形状：{result_df.shape}")
            except Exception as e:
                print(f"保存失败：{e}")
            
            
            return result_df
        
        def match_params_and_fill_min_max(top10_params_path,
                                    source_data_path,
                                    optimize_data_path,
                                    param_name_path):
            """
            1. 读取前十参数列表
            2. 匹配参数对应的序号
            3. 从源数据中提取对应序号的最小值和最大值（保留两位小数）
            4. 直接覆盖回填到原需要优化的造型数据Excel中
            :param top10_params_path: 前十参数保存路径
            :param source_data_path: 源数据文件（处理后有噪声扩充造型数据...xlsx）
            :param optimize_data_path: 需要优化的造型数据文件路径（直接覆盖此文件）
            :param param_name_path: 造型特征+技术方案名称文件路径
            """
            # 1. 读取各文件（增加异常捕获）
            try:
                # 读取前十参数
                top10_df = pd.read_excel(top10_params_path)
                top10_param_names = top10_df["参数名称"].tolist()
                print(f"\n读取到前十参数：{top10_param_names}")

                # 读取参数名称列表（获取参数序号）
                param_name_df = pd.read_excel(param_name_path, header=0)
                all_param_names = param_name_df.columns.tolist()
                print(f"总参数数量：{len(all_param_names)}")

                # 读取源数据（提取最值）
                source_df = pd.read_excel(source_data_path, header=0)
                source_cols = source_df.columns.tolist()
                print(f"源数据列数量：{len(source_cols)}")

                # 读取需要优化的造型数据（原文件）
                optimize_df = pd.read_excel(optimize_data_path, header=0)
                # 确保优化文件有“最小值”“最大值”列（无则创建）
                if "最小值" not in optimize_df.columns:
                    optimize_df["最小值"] = np.nan
                if "最大值" not in optimize_df.columns:
                    optimize_df["最大值"] = np.nan
                print(f"需要优化的造型数据列名：{optimize_df.columns.tolist()}")

            except FileNotFoundError as e:
                print(f"❌ 文件未找到：{str(e)}")
                return None
            except Exception as e:
                print(f"❌ 读取文件失败：{str(e)}")
                return None

            # 2. 匹配前十参数对应的序号并提取最值（保留两位小数）
            for param_name in top10_param_names:
                # 匹配参数在总参数列表中的序号（索引）
                if param_name in all_param_names:
                    param_col_index = all_param_names.index(param_name)
                    # 确保序号不超过源数据列数
                    if param_col_index < len(source_cols):
                        param_col_name = source_cols[param_col_index]  # 对应源数据的列名
                        print(f"\n匹配到参数：{param_name}，序号：{param_col_index}，列名：{param_col_name}")

                        # 提取源数据中该列的最小值和最大值，并保留两位小数
                        param_min = round(source_df[param_col_name].min(), 2)  # 保留两位小数
                        param_max = round(source_df[param_col_name].max(), 2)  # 保留两位小数
                        print(f"  对应最小值：{param_min}，最大值：{param_max}")

                        # 回填到优化数据中
                        if "参数名称" in optimize_df.columns:
                            # 按参数名称匹配回填（优先方案）
                            optimize_df.loc[optimize_df["参数名称"] == param_name, "最小值"] = param_min
                            optimize_df.loc[optimize_df["参数名称"] == param_name, "最大值"] = param_max
                        else:
                            # 按参数序号匹配（假设优化文件行顺序与参数序号一致）
                            param_row_index = all_param_names.index(param_name)
                            if param_row_index < len(optimize_df):
                                optimize_df.loc[param_row_index, "最小值"] = param_min
                                optimize_df.loc[param_row_index, "最大值"] = param_max
                            else:
                                print(f"警告：参数 {param_name} 序号 {param_row_index} 超出优化文件行数")
                    else:
                        print(f"警告：参数 {param_name} 序号 {param_col_index} 超出源数据列数")
                else:
                    print(f"警告：参数 {param_name} 未在总参数列表中找到，跳过")

            # 3. 直接覆盖原文件保存（核心调整：无新文件，直接写入原路径）
            try:
                # 先关闭可能占用文件的句柄，再写入
                with pd.ExcelWriter(optimize_data_path, engine='openpyxl', mode='w') as writer:
                    optimize_df.to_excel(writer, index=False)
                print(f"\n✅ 已直接覆盖原文件：{optimize_data_path}")
                print(f"✅ 最值（保留两位小数）回填完成，原文件数据已更新")
            except PermissionError:
                print(f"❌ 权限错误：无法覆盖 {optimize_data_path}，请先关闭该Excel文件")
            except Exception as e:
                print(f"❌ 覆盖文件失败：{str(e)}")
     

        #获取文件路径
        min_fre = self.current_window.ZLCB_1.currentText().strip()
        max_fre = self.current_window.ZLCB_2.currentText().strip()
        
        try:
            model_path=self.current_window.ZL_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            newinput_file_path=self.current_window.ZL_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择进行灵敏度排序的数据文件！")
        
        MIV, IV1, IV2 = MIV_calculate.calculate_result(self.input_file_path, self.output_file_path, newinput_file_path, model_path, self.Characteristic_name,self.huancun)

        freq_labels = ["200Hz", "250Hz", "315Hz", "400Hz", "500Hz", "630Hz", "800Hz", "1000Hz", "1250Hz",
                   "1600Hz", "2000Hz", "2500Hz", "3150Hz", "4000Hz", "5000Hz", "6300Hz", "8000Hz"]
        fre_index1 = freq_labels.index(max_fre)
        fre_index2 = freq_labels.index(min_fre)
        if fre_index1 < fre_index2:
            QMessageBox.warning(self.current_window, "输入错误", "分析频率范围最小值不能大于最大值！")
        elif fre_index1 == fre_index2:
            miv_for_freq1 = IV1[:,fre_index1]
            miv_for_freq2 = IV2[:,fre_index1]
            freq_title = f'风噪 {freq_labels[fre_index1]} 灵敏度分析'
            plot_sensitivity(miv_for_freq1, miv_for_freq2, self.Characteristic_name, freq_title, "ZLwidget", save_path=self.huancun)
        else:
            miv_data = MIV[:,fre_index2:fre_index1+1]
            freq_title = f'风噪 {freq_labels[fre_index2]}Hz-{freq_labels[fre_index1]}Hz 灵敏度分析'
            y_label = freq_labels[fre_index2:fre_index1+1]
            plot_sensitivityonly(miv_data, self.Characteristic_name, y_label, freq_title, "ZLwidget", save_path=self.huancun)
        
        #生成优化方案初始文件
        name = pd.read_excel(self.all_characteristic, header=0)#获取技术方案名称
        param_names = name.columns.tolist()  # 第一行作为列名
        new_input_data = pd.read_excel(newinput_file_path)
        data = new_input_data.iloc[0, :].values
        new_data = data.T
        # 定义列名
        #columns = ["参数名称", "原始值", "最小值", "最大值"]
        df = pd.DataFrame(param_names, columns=['参数名称'])
        df['原始值'] = new_data
        df['最小值'] = new_data
        df['最大值'] = new_data
        save_path = os.path.join(self.huancun, "优化方案.xlsx")
        df.to_excel(save_path, index=False, engine="openpyxl")
        # 输出优化方案
        # 1. 先执行MIV总和排序
        name = pd.read_excel(self.Characteristic_name, header=0)#获取技术方案名称
        labels = name.columns.tolist()  # 第一行作为列名
        
        all_rank_save_path = os.path.join(self.huancun, "全频段参数MIV_前十.xlsx")
        os.makedirs(os.path.dirname(all_rank_save_path), exist_ok=True)
        sum_rank_save_path = os.path.join(self.huancun, "全频段参数MIV总和_前十.xlsx")
        os.makedirs(os.path.dirname(all_rank_save_path), exist_ok=True)
        param_rank_result = sum_and_rank_params_from_heatmap(
            MIV=MIV,
            param_labels=labels,
            freq_labels=freq_labels,
            save_path=sum_rank_save_path
        )
        result_df = all_params_from_heatmap(MIV, labels, freq_labels, all_rank_save_path)
        plot_excel_table_widget(all_rank_save_path, "Sheet1", "ZLwidget2")

        # 2. 再执行最值回填（直接覆盖原文件，最值保留两位小数）
        # 定义各文件路径
        top10_params_path = sum_rank_save_path
        source_data_path = self.input_file_path
        optimize_data_path = save_path  # 原文件路径（直接覆盖）
        param_name_path = self.Characteristic_name

        # 执行回填（无额外保存路径，直接覆盖原文件）
        match_params_and_fill_min_max(
            top10_params_path=top10_params_path,
            source_data_path=source_data_path,
            optimize_data_path=optimize_data_path,
            param_name_path=param_name_path
        )
 
    def save_results(self):
        """在灵敏度分析完成后保存分析结果"""
        if not hasattr(self, "model_train"):
            print("❌ 尚未进行模型训练，无法保存！")
            return

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存分析结果", "", "文件夹 (*)")
        try:
            # 4. 创建新文件夹（exist_ok=False 避免重名）
            os.makedirs(save_path, exist_ok=False)
        except FileExistsError:
            QMessageBox.critical(None, "错误", f"文件夹「{save_path}」已存在！")
            return
        except Exception as e:
            QMessageBox.critical(None, "错误", f"创建文件夹失败：{str(e)}")
            return

        #设置要移动文件的路径
        MIV_path = os.path.join(self.huancun, "MIV数组.xlsx")
        IV1_path = os.path.join(self.huancun, "IV1数组.xlsx")
        IV2_path = os.path.join(self.huancun, "IV2数组.xlsx")
        heatmap_path = os.path.join(self.huancun, "风噪灵敏度分析结果.png")
        Optim_result_path = os.path.join(self.huancun, "优化方案.xlsx")

        # 5. 检查要移动的模型是否存在
        if not os.path.exists(MIV_path):
            QMessageBox.critical(None, "错误", f"指定文件MIV数组.xlsx不存在！")
            return
        if not os.path.exists(IV1_path):
            QMessageBox.critical(None, "错误", f"指定文件IV1数组.xlsx不存在！")
            return
        if not os.path.exists(IV2_path):
            QMessageBox.critical(None, "错误", f"指定文件IV2数组.xlsx不存在！")
            return
        if not os.path.exists(heatmap_path):
            QMessageBox.critical(None, "错误", f"指定文件风噪灵敏度分析结果.png不存在！")
            return
        if not os.path.exists(Optim_result_path):
            QMessageBox.critical(None, "错误", f"指定文件优化方案.xlsx不存在！")
            return


        # 6. 拼接文件移动后的新路径
        new_MIV_path = os.path.join(save_path, "MIV数组.xlsx") #保存MIV数组
        new_IV1_path = os.path.join(save_path, "IV1数组.xlsx") #保存IV1数组
        new_IV2_path = os.path.join(save_path, "IV2数组.xlsx") #保存IV2数组
        new_heatmap_path = os.path.join(save_path, "风噪灵敏度分析结果.png") #保存热力图
        new_Optim_result_path = os.path.join(save_path, "优化方案.xlsx") #保存优化方案

        try:
            # 7. 移动文件到新文件夹
            shutil.move(MIV_path, new_MIV_path)
            shutil.move(IV1_path, new_IV1_path)
            shutil.move(IV2_path, new_IV2_path)
            shutil.move(heatmap_path, new_heatmap_path)
            shutil.move(Optim_result_path, new_Optim_result_path)
            self.current_window.ZJP_2.setText(new_Optim_result_path)
            self.current_window.ZJX_2.setText(new_Optim_result_path)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"移动文件失败：{str(e)}")
            return

        # 8. 弹窗提示文件保存的路径
        QMessageBox.information(
            None, "成功", f"文件已移动至：\n{save_path}"
        )               
        

    
    #----基于具体频段-----
    def select_folder_pinduan(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path

        if hasattr(self.current_window, "ZJP_1"):
            self.current_window.ZJP_1.setText(pth_path)

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)

    def select_file_zxyh_pinduan(self):
        """选择 new_input_path 文件并自动读取原始值、最小值、最大值，填入 lineEdit"""
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "选择需要优化的造型数据",
            "",
            "Excel 文件 (*.xlsx)"
        )

        if not file_path:
            return

        # 写入 ZJP_2
        self.current_window.ZJP_2.setText(file_path)

        # ---------------------- 读取 Excel 并自动填入界面 ---------------------- #
        try:
            import pandas as pd

            df = pd.read_excel(file_path, sheet_name="Sheet1")

            required_cols = ["原始值", "最小值", "最大值"]
            if not all(col in df.columns for col in required_cols):
                QMessageBox.warning(
                    None, "格式错误",
                    "Excel sheet1 必须包含 '原始值'、'最小值'、'最大值' 三列！"
                )
                return

            base_params = df['原始值'].values
            param_min = df['最小值'].values
            param_max = df['最大值'].values

            # 转换为原生 python float，避免 np.float64(...) 的字符串
            try:
                param_min_py = [float(x) for x in param_min]
                param_max_py = [float(x) for x in param_max]
                base_params_py = [float(x) for x in base_params]
            except Exception:
                # 如果逐元素转换失败，退回到逐项用 safe 提取
                param_min_py = [self._safe_to_float(str(x)) for x in param_min]
                param_max_py = [self._safe_to_float(str(x)) for x in param_max]
                base_params_py = [self._safe_to_float(str(x)) for x in base_params]

            # 自动识别可调整参数
            adjust_indices = [i for i in range(len(base_params_py)) if param_min_py[i] != param_max_py[i]]

            QMessageBox.information(
                None, "读取成功",
                "已成功读取 Excel：\n"
                f"识别到可调整参数个数：{len(adjust_indices)}"
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(None, "错误", f"读取 Excel 时出错：\n{e}")

    def plot_moxingyouhua_pinduan(self):
        """绘制模型预测结果图"""

        # 可视化原始与优化方案结果对比
        def visualize_freq_comparison(original, optimized, target_indices, widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制原始与优化方案的频点对比折线图，高亮显示目标频段。
            支持动态适应widget尺寸，并可选保存图像。
            """
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 频率标签（根据数据长度自动截取）
            freq_labels = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                                2000, 2500, 3150, 4000, 5000, 6300, 8000][:len(original)]
            x = np.arange(len(freq_labels))

            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（约100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 绘制折线
            ax.plot(x, original, 'ro-', linewidth=2, markersize=6, label='原始方案')
            ax.plot(x, optimized, 'bo-', linewidth=2, markersize=6, label='优化方案')

            # 高亮目标频段
            if target_indices:
                target_x = np.array(target_indices)
                ax.fill_between(target_x, original[target_x], optimized[target_x],
                                color='green', alpha=0.3, label='优化目标频段')

            # 设置坐标轴和标题
            ax.set_xticks(x)
            ax.set_xticklabels(freq_labels, rotation=45, fontsize=12)
            ax.set_xlabel('频率(Hz)', fontsize=14)
            ax.set_ylabel('噪声值(dB)', fontsize=14)
            ax.set_title('原始方案与优化方案的频点对比', fontsize=16, pad=15)
            ax.legend(fontsize=12)
            ax.grid(alpha=0.3)
            
            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, '频点对比折线图.png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
                print(f"频点对比折线图已保存至: {save_pathnew}")

        # 可视化调整参数的前后
        def visualize_param_changes(original_params, optimized_params, adjust_indices, 
                                    param_min_dict, param_max_dict, widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制调整参数的前后对比柱状图，并显示每个参数的调整范围。
            支持动态适应widget尺寸，并可选保存图像。
            """
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False

            param_indices = adjust_indices
            original_values = [original_params[i] for i in param_indices]
            optimized_values = [optimized_params[i] for i in param_indices]
            param_ranges = [f"{param_min_dict[i]}-{param_max_dict[i]}" for i in param_indices]

            x = np.arange(len(param_indices))
            width = 0.35

            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（约100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 绘制柱状图
            ax.bar(x - width / 2, original_values, width, label='原始参数值', alpha=0.8, color='#ff7f0e')
            ax.bar(x + width / 2, optimized_values, width, label='优化参数值', alpha=0.8, color='#1f77b4')

            # 设置x轴标签：参数索引 + 换行 + 调整范围
            x_labels = [f'参数{i}\n({r})' for i, r in zip(param_indices, param_ranges)]
            ax.set_xticks(x)
            ax.set_xticklabels(x_labels, rotation=45, fontsize=11, ha='center')
            
            # 设置标题和轴标签
            ax.set_xlabel('参数索引及调整范围', fontsize=14)
            ax.set_ylabel('参数值', fontsize=14)
            ax.set_title('调整参数的前后对比', fontsize=16, pad=15)
            ax.legend(fontsize=12)
            ax.grid(axis='y', alpha=0.3)

            # 在柱子上方标注数值
            max_val = max(max(original_values), max(optimized_values))
            offset = max_val * 0.02  # 略微上移，避免重叠
            for i, (orig, opt) in enumerate(zip(original_values, optimized_values)):
                ax.text(i - width / 2, orig + offset, f'{orig:.2f}', ha='center', fontsize=10, fontweight='bold', rotation=45)
                ax.text(i + width / 2, opt + offset, f'{opt:.2f}', ha='center', fontsize=10, fontweight='bold', rotation=45)

            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, '参数调整前后对比柱状图.png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
                print(f"参数对比图已保存至: {save_pathnew}")  
        
        #进行优化
        try:
            model_path=self.current_window.ZJP_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            new_input_path=self.current_window.ZJP_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择进行灵敏度排序的数据文件！")
        
        input_file_path = self.input_file_path #输入归一化
        output_file_path = self.output_file_path #输出归一化
        result_save_path = os.path.join(self.huancun, f'参数优化结果.xlsx') #优化结果保存路径
        full_freq_table_path = os.path.join(self.huancun, f'噪声值对比表.xlsx') #200-8000Hz噪声值对比表保存路径
        target_freq_min = int(self.current_window.ZJPCB_1.currentText())
        target_freq_max = int(self.current_window.ZJPCB_2.currentText())
        try:
            generations = int(self.current_window.ZJP_7.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法迭代次数必须为数字！")
        try:
            pop_size = int(self.current_window.ZJP_6.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法方案数量必须为数字！")
        #计算结果    
        original_freq_values, best_freq_values, target_indices, base_params, best_params, adjust_indices, param_min_dict, param_max_dict = optimization_pinduan.optimization_program(model_path, input_file_path, output_file_path, new_input_path, result_save_path, full_freq_table_path, target_freq_min, target_freq_max, pop_size, generations)
        #生成频点对比折线图
        visualize_freq_comparison(original_freq_values, best_freq_values, target_indices, 'ZJPwidget_1', self.huancun)
        #生成参数调整对比柱状图
        visualize_param_changes(base_params, best_params, adjust_indices, param_min_dict, param_max_dict, 'ZJPwidget_2', self.huancun)
    #----保存结果----
    def save_result_pinduan(self):
        """保存参数优化结果"""

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存分析结果", "", "文件夹 (*)")
        try:
            # 4. 创建新文件夹（exist_ok=False 避免重名）
            os.makedirs(save_path, exist_ok=False)
        except FileExistsError:
            QMessageBox.critical(None, "错误", f"文件夹「{save_path}」已存在！")
            return
        except Exception as e:
            QMessageBox.critical(None, "错误", f"创建文件夹失败：{str(e)}")
            return

        #设置要移动文件的路径

        result_path = os.path.join(self.huancun, "参数优化结果.xlsx")
        full_freq_table_path = os.path.join(self.huancun, "噪声值对比表.xlsx")
        zhexian_path = os.path.join(self.huancun, "频点对比折线图.png")
        zhuzhuang_path = os.path.join(self.huancun, "参数调整前后对比柱状图.png")

        # 5. 检查要移动的模型是否存在
        if not os.path.exists(result_path):
            QMessageBox.critical(None, "错误", f"指定文件参数优化结果.xlsx不存在！")
            return
        if not os.path.exists(full_freq_table_path):
            QMessageBox.critical(None, "错误", f"指定文件噪声值对比表.xlsx不存在！")
            return
        if not os.path.exists(zhexian_path):
            QMessageBox.critical(None, "错误", f"指定文件频点对比折线图.png不存在！")
            return
        if not os.path.exists(zhuzhuang_path):
            QMessageBox.critical(None, "错误", f"指定文件参数调整前后对比柱状图.png不存在！")
            return



        # 6. 拼接文件移动后的新路径
        new_result_path = os.path.join(save_path, "参数优化结果.xlsx") #保存优化结果
        new_full_freq_table_path = os.path.join(save_path, "噪声值对比表.xlsx") #保存噪声曲线对比
        new_zhexian_path = os.path.join(save_path, "频点对比折线图.png") #保存频点对比折线图
        new_zhuzhuang_path = os.path.join(save_path, "参数调整前后对比柱状图.png") #保存参数对比柱状图


        try:
            # 7. 移动文件到新文件夹
            shutil.move(result_path, new_result_path)
            shutil.move(full_freq_table_path, new_full_freq_table_path)
            shutil.move(zhexian_path, new_zhexian_path)
            shutil.move(zhuzhuang_path, new_zhuzhuang_path)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"移动文件失败：{str(e)}")
            return

        # 8. 弹窗提示文件保存的路径
        QMessageBox.information(
            None, "成功", f"文件已移动至：\n{save_path}"
        )
    
    #----基于整体响度-----
    def select_folder_xiangdu(self):
        """选择文件夹，自动搜索 .pth、输入数据.xlsx、输出数据.xlsx 并写入相应输入框"""
        folder_path = QFileDialog.getExistingDirectory(None, "选择包含模型和数据的文件夹")
        if not folder_path:
            return

        pth_path = ""

        for file_name in os.listdir(folder_path):
            lower_name = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if lower_name.endswith(".pth") and not pth_path:
                pth_path = full_path
            elif file_name == "输入数据.xlsx":
                input_xlsx_path = full_path
            elif file_name == "输出数据.xlsx":
                output_xlsx_path = full_path

        if hasattr(self.current_window, "ZJX_1"):
            self.current_window.ZJX_1.setText(pth_path)

        msg = f"📁 已选择文件夹：{folder_path}\n"
        msg += f"\n模型文件 (.pth)：{pth_path if pth_path else '未找到'}"
        QMessageBox.information(None, "文件检测结果", msg)

    def select_file_zxyh_xiangdu(self):
        """选择 new_input_path 文件并自动读取原始值、最小值、最大值，填入 lineEdit"""
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "选择需要优化的造型数据",
            "",
            "Excel 文件 (*.xlsx)"
        )

        if not file_path:
            return

        # 写入 ZJP_2
        self.current_window.ZJX_2.setText(file_path)

        # ---------------------- 读取 Excel 并自动填入界面 ---------------------- #
        try:
            import pandas as pd

            df = pd.read_excel(file_path, sheet_name="Sheet1")

            required_cols = ["原始值", "最小值", "最大值"]
            if not all(col in df.columns for col in required_cols):
                QMessageBox.warning(
                    None, "格式错误",
                    "Excel sheet1 必须包含 '原始值'、'最小值'、'最大值' 三列！"
                )
                return

            base_params = df['原始值'].values
            param_min = df['最小值'].values
            param_max = df['最大值'].values

            # 转换为原生 python float，避免 np.float64(...) 的字符串
            try:
                param_min_py = [float(x) for x in param_min]
                param_max_py = [float(x) for x in param_max]
                base_params_py = [float(x) for x in base_params]
            except Exception:
                # 如果逐元素转换失败，退回到逐项用 safe 提取
                param_min_py = [self._safe_to_float(str(x)) for x in param_min]
                param_max_py = [self._safe_to_float(str(x)) for x in param_max]
                base_params_py = [self._safe_to_float(str(x)) for x in base_params]

            # 自动识别可调整参数
            adjust_indices = [i for i in range(len(base_params_py)) if param_min_py[i] != param_max_py[i]]

            QMessageBox.information(
                None, "读取成功",
                "已成功读取 Excel：\n"
                f"识别到可调整参数个数：{len(adjust_indices)}"
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(None, "错误", f"读取 Excel 时出错：\n{e}")

    def plot_moxingyouhua_xiangdu(self):
        """绘制模型预测结果图"""

        # 可视化原始与优化方案结果对比
        def visualize_freq_comparison(original, optimized, original_loudness, optimized_loudness,
                                    widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制原始与优化方案的频点对比折线图，
            图例中显示各自响度值（sone），支持动态适应widget尺寸，并可选保存图像。
            """
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False

            # 频率标签（根据数据长度自动截取）
            freq_labels = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600,
                        2000, 2500, 3150, 4000, 5000, 6300, 8000][:len(original)]
            x = np.arange(len(freq_labels))

            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（约100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 绘制折线，并在图例中显示响度
            ax.plot(x, original, 'ro-', linewidth=2, markersize=6,
                    label=f'原始方案 (响度: {original_loudness:.2f} sone)')
            ax.plot(x, optimized, 'bo-', linewidth=2, markersize=6,
                    label=f'优化方案 (响度: {optimized_loudness:.2f} sone)')

            # 设置坐标轴
            ax.set_xticks(x)
            ax.set_xticklabels(freq_labels, rotation=45, fontsize=11)
            ax.set_xlabel('频率(Hz)', fontsize=14)
            ax.set_ylabel('噪声值(dB)', fontsize=14)
            ax.set_title('原始方案与优化方案的频点对比（响度优化）', fontsize=16, pad=15)
            ax.legend(fontsize=12, loc='upper right')
            ax.grid(alpha=0.3)

            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, '频点对比折线图(响度).png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
                print(f"频点对比折线图（含响度）已保存至: {save_pathnew}")

        # 可视化调整参数的前后
        def visualize_param_changes(original_params, optimized_params, adjust_indices, 
                                    param_min_dict, param_max_dict, widget_name, save_path=None):
            """
            在指定的UI QWidget中绘制调整参数的前后对比柱状图，
            x轴显示参数索引及调整范围，并在柱子上方标注数值。
            支持动态适应widget尺寸，并可选保存图像。
            """
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei']
            plt.rcParams['axes.unicode_minus'] = False

            param_indices = adjust_indices
            original_values = [original_params[i] for i in param_indices]
            optimized_values = [optimized_params[i] for i in param_indices]
            param_ranges = [f"{param_min_dict.get(i, '?')}-{param_max_dict.get(i, '?')}" for i in param_indices]

            x = np.arange(len(param_indices))
            width = 0.35

            # 获取指定的QWidget
            plot_widget = self.current_window.findChild(QWidget, widget_name)
            if not plot_widget:
                print(f"警告: 找不到名为'{widget_name}'的QWidget")
                return
            
            # 获取widget的宽度和高度（单位：像素）
            widget_width = plot_widget.width()
            widget_height = plot_widget.height()
            
            # 创建matplotlib图形，尺寸转换为英寸（约100 dpi）
            fig, ax = plt.subplots(figsize=(widget_width / 100, widget_height / 100))
            
            # 绘制并列柱状图
            ax.bar(x - width / 2, original_values, width, label='原始参数值', alpha=0.8, color='#ff7f0e')
            ax.bar(x + width / 2, optimized_values, width, label='优化参数值', alpha=0.8, color='#1f77b4')

            # x轴标签：参数索引 + 换行 + 调整范围
            x_labels = [f'参数{i}\n({r})' for i, r in zip(param_indices, param_ranges)]
            ax.set_xticks(x)
            ax.set_xticklabels(x_labels, rotation=45, fontsize=11, ha='center')

            # 设置标题和轴标签
            ax.set_xlabel('参数索引及调整范围', fontsize=14)
            ax.set_ylabel('参数值', fontsize=14)
            ax.set_title('调整参数的前后对比', fontsize=16, pad=15)
            ax.legend(fontsize=12)
            ax.grid(axis='y', alpha=0.3)

            # 在每个柱子上方标注数值
            max_val = max(max(original_values or [0]), max(optimized_values or [0]))
            offset = max_val * 0.02 if max_val > 0 else 0.1  # 避免全零时重叠
            for i, (orig, opt) in enumerate(zip(original_values, optimized_values)):
                ax.text(i - width / 2, orig + offset, f'{orig:.2f}', 
                        ha='center', va='bottom', fontsize=10, fontweight='bold', rotation=45)
                ax.text(i + width / 2, opt + offset, f'{opt:.2f}', 
                        ha='center', va='bottom', fontsize=10, fontweight='bold', rotation=45)

            # 布局调整
            plt.tight_layout()
            
            # 将图表嵌入到QWidget中
            canvas = FigureCanvas(fig)
            canvas.setParent(plot_widget)
            canvas.draw()
            
            # 适应widget大小
            canvas.setGeometry(plot_widget.rect())
            canvas.setSizePolicy(plot_widget.sizePolicy())
            
            # 添加到布局（如果尚未有布局，则创建）
            layout = plot_widget.layout()
            if layout is None:
                layout = QVBoxLayout(plot_widget)
            
            # 清除可能已存在的旧canvas（避免重复叠加）
            for i in reversed(range(layout.count())):
                old_widget = layout.itemAt(i).widget()
                if isinstance(old_widget, FigureCanvas):
                    old_widget.deleteLater()
            
            layout.addWidget(canvas)
            
            # 可选：保存图像到文件
            if save_path:
                save_pathnew = os.path.join(save_path, '参数调整前后对比柱状图(响度).png')
                fig.savefig(save_pathnew, dpi=300, bbox_inches='tight')
                print(f"参数对比图已保存至: {save_pathnew}")  
        
        #进行优化
        try:
            model_path=self.current_window.ZJX_1.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择模型文件！")
        try:
            new_input_path=self.current_window.ZJX_2.text().strip()
        except ValueError:
            QMessageBox.warning(self.current_window, "缺少必要的输入", "请选择进行优化的数据文件！")
        
        input_file_path = self.input_file_path #输入归一化
        output_file_path = self.output_file_path #输出归一化
        result_save_path = os.path.join(self.huancun, f'参数优化结果_响度.xlsx') #优化结果保存路径

        try:
            target_loudness = float(self.current_window.ZJX_8.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "响度值必须为数字！")
            
        try:
            generations = int(self.current_window.ZJX_7.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法迭代次数必须为数字！")
        try:
            pop_size = int(self.current_window.ZJX_6.text().strip())
        except ValueError:
            QMessageBox.warning(self.current_window, "输入错误", "遗传算法方案数量必须为数字！")
        #计算结果    
        original_freq_values, best_freq_values, original_loudness, best_loudness, base_params, best_params, adjust_indices, param_min_dict, param_max_dict,error = optimization_xiangdu.optimization_program(model_path, input_file_path, output_file_path, new_input_path, result_save_path, target_loudness, pop_size, generations)
        if error ==1:
            QMessageBox.warning(self.current_window, "优化失败", "未找到低于目标响度的方案，返回原始方案！")
        #生成频点对比折线图
        visualize_freq_comparison(original_freq_values, best_freq_values, original_loudness, best_loudness, 'ZJXwidget_1', self.huancun)
        #生成参数调整对比柱状图
        visualize_param_changes(base_params, best_params, adjust_indices, param_min_dict, param_max_dict, 'ZJXwidget_2', self.huancun)

    #----保存结果----
    def save_result_xiangdu(self):
        """保存参数优化结果"""

        # 弹出文件选择对话框
        save_path, _ = QFileDialog.getSaveFileName(self.current_window, "保存分析结果", "", "文件夹 (*)")
        try:
            # 4. 创建新文件夹（exist_ok=False 避免重名）
            os.makedirs(save_path, exist_ok=False)
        except FileExistsError:
            QMessageBox.critical(None, "错误", f"文件夹「{save_path}」已存在！")
            return
        except Exception as e:
            QMessageBox.critical(None, "错误", f"创建文件夹失败：{str(e)}")
            return

        #设置要移动文件的路径

        result_path = os.path.join(self.huancun, "参数优化结果_响度.xlsx")
        zhexian_path = os.path.join(self.huancun, "频点对比折线图(响度).png")
        zhuzhuang_path = os.path.join(self.huancun, "参数调整前后对比柱状图(响度).png")

        # 5. 检查要移动的模型是否存在
        if not os.path.exists(result_path):
            QMessageBox.critical(None, "错误", f"指定文件参数优化结果.xlsx不存在！")
            return
        if not os.path.exists(zhexian_path):
            QMessageBox.critical(None, "错误", f"频点对比折线图(响度).png不存在！")
            return
        if not os.path.exists(zhuzhuang_path):
            QMessageBox.critical(None, "错误", f"指定文件参数调整前后对比柱状图(响度).png不存在！")
            return



        # 6. 拼接文件移动后的新路径
        new_result_path = os.path.join(save_path, "参数优化结果_响度.xlsx") #保存优化结果
        new_zhexian_path = os.path.join(save_path, "频点对比折线图(响度).png") #保存噪声对比曲线
        new_zhuzhuang_path = os.path.join(save_path, "参数调整前后对比柱状图(响度).png") #保存参数对比柱状图
 

        try:
            # 7. 移动文件到新文件夹
            shutil.move(result_path, new_result_path)
            shutil.move(zhexian_path, new_zhexian_path)
            shutil.move(zhuzhuang_path, new_zhuzhuang_path)
        except Exception as e:
            QMessageBox.critical(None, "错误", f"移动文件失败：{str(e)}")
            return

        # 8. 弹窗提示文件保存的路径
        QMessageBox.information(
            None, "成功", f"文件已移动至：\n{save_path}"
        )
  




if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    sys.exit(app.exec())